#!/usr/bin/env python3
"""
Create a test Excel file with formulas for testing the CSV Editor integration
"""

import openpyxl
from openpyxl.styles import Font, PatternFill

def create_test_excel():
    """Create an Excel file with various formulas for testing"""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formula Test"
    
    # Add headers
    headers = ["Product", "Price", "Quantity", "Total", "Tax Rate", "Tax Amount", "Final Price"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Add sample data
    data = [
        ["Laptop", 1000, 2],
        ["Mouse", 25, 5],
        ["Keyboard", 75, 3],
        ["Monitor", 300, 1],
        ["Headphones", 150, 2]
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Add tax rate (10%)
    for row in range(2, 7):
        ws.cell(row=row, column=5, value=0.1)
    
    # Add formulas
    for row in range(2, 7):
        # Total = Price * Quantity (column D = B * C)
        ws.cell(row=row, column=4, value=f"=B{row}*C{row}")
        
        # Tax Amount = Total * Tax Rate (column F = D * E)
        ws.cell(row=row, column=6, value=f"=D{row}*E{row}")
        
        # Final Price = Total + Tax Amount (column G = D + F)
        ws.cell(row=row, column=7, value=f"=D{row}+F{row}")
    
    # Add summary row with aggregate functions
    summary_row = 8
    ws.cell(row=summary_row, column=1, value="SUMMARY")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    
    # Add aggregate formulas
    ws.cell(row=summary_row, column=2, value="=AVERAGE(B2:B6)")  # Average price
    ws.cell(row=summary_row, column=3, value="=SUM(C2:C6)")      # Total quantity
    ws.cell(row=summary_row, column=4, value="=SUM(D2:D6)")      # Total amount
    ws.cell(row=summary_row, column=6, value="=SUM(F2:F6)")      # Total tax
    ws.cell(row=summary_row, column=7, value="=SUM(G2:G6)")      # Grand total
    
    # Add conditional formulas
    condition_row = 10
    ws.cell(row=condition_row, column=1, value="High Value?")
    ws.cell(row=condition_row, column=2, value='=IF(G2>1000,"YES","NO")')  # Check if first item is high value
    
    ws.cell(row=condition_row+1, column=1, value="Count Expensive")
    ws.cell(row=condition_row+1, column=2, value='=COUNTIF(B2:B6,">100")')  # Count items over $100
    
    ws.cell(row=condition_row+2, column=1, value="Max Price")
    ws.cell(row=condition_row+2, column=2, value="=MAX(B2:B6)")  # Maximum price
    
    ws.cell(row=condition_row+3, column=1, value="Min Price")
    ws.cell(row=condition_row+3, column=2, value="=MIN(B2:B6)")  # Minimum price
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    filename = "test_formulas_integration.xlsx"
    wb.save(filename)
    print(f"Test Excel file created: {filename}")
    print("\nFile contains:")
    print("- Basic arithmetic formulas (multiplication, addition)")
    print("- SUM, AVERAGE, COUNT, MAX, MIN functions")
    print("- IF and COUNTIF conditional functions")
    print("- Cell references and ranges")
    print("\nUse this file to test the CSV Editor formula integration.")
    
    return filename

if __name__ == "__main__":
    create_test_excel()