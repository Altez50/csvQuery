#!/usr/bin/env python3
"""
Test script for XLSX saving functionality in CSV Query Tool.
This script creates sample data, loads it into the CSV editor,
and demonstrates the new XLSX export feature.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
import os

def create_test_data_csv():
    """Create a test CSV file with varied data types"""
    
    data = {
        'ID': [1, 2, 3, 4, 5],
        'Product': ['Laptop', 'Mouse', 'Keyboard', 'Monitor', 'Headphones'],
        'Category': ['Electronics', 'Accessories', 'Accessories', 'Electronics', 'Audio'],
        'Price': [999.99, 25.50, 75.00, 299.99, 149.99],
        'Stock': [15, 100, 45, 8, 23],
        'Description': [
            'High-performance laptop with 16GB RAM',
            'Wireless optical mouse with ergonomic design',
            'Mechanical keyboard with RGB backlighting',
            '27-inch 4K monitor with HDR support',
            'Noise-cancelling wireless headphones'
        ],
        'Available': ['Yes', 'Yes', 'No', 'Yes', 'Yes'],
        'Rating': [4.5, 4.2, 4.8, 4.6, 4.3]
    }
    
    df = pd.DataFrame(data)
    csv_file = 'test_data_for_xlsx.csv'
    df.to_csv(csv_file, index=False)
    
    print(f"✅ Created test CSV file: {csv_file}")
    print(f"   📊 Data: {len(df)} rows, {len(df.columns)} columns")
    return csv_file

def create_formatted_excel_for_comparison():
    """Create a formatted Excel file to compare with exported XLSX"""
    
    data = {
        'ID': [1, 2, 3, 4, 5],
        'Product': ['Laptop', 'Mouse', 'Keyboard', 'Monitor', 'Headphones'],
        'Category': ['Electronics', 'Accessories', 'Accessories', 'Electronics', 'Audio'],
        'Price': [999.99, 25.50, 75.00, 299.99, 149.99],
        'Stock': [15, 100, 45, 8, 23],
        'Description': [
            'High-performance laptop with 16GB RAM',
            'Wireless optical mouse with ergonomic design',
            'Mechanical keyboard with RGB backlighting',
            '27-inch 4K monitor with HDR support',
            'Noise-cancelling wireless headphones'
        ],
        'Available': ['Yes', 'Yes', 'No', 'Yes', 'Yes'],
        'Rating': [4.5, 4.2, 4.8, 4.6, 4.3]
    }
    
    df = pd.DataFrame(data)
    excel_file = 'formatted_comparison.xlsx'
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Products', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Products']
        
        # Header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
        
        # Data formatting
        # Price column - green background for high values
        price_col = df.columns.get_loc('Price') + 1
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=price_col)
            if cell.value and float(cell.value) > 200:
                cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        
        # Stock column - red background for low stock
        stock_col = df.columns.get_loc('Stock') + 1
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=stock_col)
            if cell.value and int(cell.value) < 20:
                cell.fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
        
        # Available column - different colors for Yes/No
        available_col = df.columns.get_loc('Available') + 1
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=available_col)
            if cell.value == 'Yes':
                cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            elif cell.value == 'No':
                cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
        
        # Set column widths
        column_widths = {
            'A': 8,   # ID
            'B': 15,  # Product
            'C': 12,  # Category
            'D': 10,  # Price
            'E': 8,   # Stock
            'F': 40,  # Description
            'G': 12,  # Available
            'H': 10   # Rating
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
    
    print(f"✅ Created formatted Excel file for comparison: {excel_file}")
    return excel_file

def analyze_xlsx_file(xlsx_file):
    """Analyze an XLSX file to show its formatting and structure"""
    print(f"\n🔍 Analyzing XLSX file: {xlsx_file}")
    
    try:
        wb = openpyxl.load_workbook(xlsx_file)
        ws = wb.active
        
        print(f"   📋 Worksheet: {ws.title}")
        print(f"   📊 Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        
        # Check headers
        print(f"   📝 Headers:")
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            header_name = cell.value
            has_formatting = (
                cell.fill.start_color.rgb != '00000000' or 
                (cell.font.color and cell.font.color.rgb != '00000000') or
                cell.font.bold
            )
            print(f"      Column {col}: '{header_name}' (Formatted: {has_formatting})")
        
        # Check column widths
        print(f"   📏 Column widths:")
        for col_letter, column_dimension in ws.column_dimensions.items():
            width = column_dimension.width if column_dimension.width else "Auto"
            print(f"      {col_letter}: {width}")
        
        # Sample data formatting
        print(f"   🎨 Sample formatting (first 3 data rows):")
        for row in range(2, min(5, ws.max_row + 1)):
            formatted_cells = 0
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if (cell.fill.start_color.rgb != '00000000' or 
                    (cell.font.color and cell.font.color.rgb != '00000000')):
                    formatted_cells += 1
            print(f"      Row {row}: {formatted_cells}/{ws.max_column} cells have formatting")
        
    except Exception as e:
        print(f"❌ Error analyzing XLSX file: {e}")

def main():
    """Main test function"""
    print("🧪 Testing XLSX Saving Functionality")
    print("=" * 50)
    
    # Create test files
    csv_file = create_test_data_csv()
    comparison_file = create_formatted_excel_for_comparison()
    
    # Analyze the comparison file
    analyze_xlsx_file(comparison_file)
    
    print("\n✅ Test files created successfully!")
    print("\n📝 Manual Testing Instructions:")
    print("   1. Open the CSV Query Tool application")
    print(f"   2. Load the CSV file: {csv_file}")
    print("      - Use 'Load CSV' button")
    print("      - Verify data is displayed correctly")
    print("   3. Test XLSX saving:")
    print("      - Click '📊 Save XLSX' button")
    print("      - Choose a filename (e.g., 'exported_test.xlsx')")
    print("      - Verify the file is saved successfully")
    print("   4. Load a formatted Excel file:")
    print(f"      - Load the formatted file: {comparison_file}")
    print("      - Use 'Load Excel' with formatting options enabled")
    print("      - Observe the formatting preservation")
    print("   5. Save the formatted data as XLSX:")
    print("      - Click '📊 Save XLSX' button")
    print("      - Save as 'exported_formatted.xlsx'")
    print("      - Compare with original formatting")
    
    print("\n🎯 Expected Results:")
    print("   📄 CSV → XLSX: Clean export with proper headers and auto-sized columns")
    print("   🎨 Formatted Excel → XLSX: Formatting preservation including:")
    print("      - Header styling (blue background, white text, bold)")
    print("      - Cell background colors")
    print("      - Column widths")
    print("      - Font properties")
    print("   📊 File compatibility: Generated XLSX files should open in Excel/LibreOffice")
    
    print("\n🔧 Features to Test:")
    print("   ✓ Basic XLSX export from CSV data")
    print("   ✓ Formatting preservation from Excel imports")
    print("   ✓ Column width preservation")
    print("   ✓ Header styling")
    print("   ✓ Cell background and text colors")
    print("   ✓ Font properties (family, size, bold, italic)")
    print("   ✓ Auto-sizing for columns without explicit widths")
    print("   ✓ Error handling for missing openpyxl library")
    
    print("\n💡 Additional Tests:")
    print("   - Try editing data in the CSV editor and then saving to XLSX")
    print("   - Test with empty cells and special characters")
    print("   - Verify large datasets export correctly")
    print("   - Check file size and performance")

if __name__ == "__main__":
    main()