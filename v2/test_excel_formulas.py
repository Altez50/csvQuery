#!/usr/bin/env python3
"""
Тестирование поддержки формул Excel
Создание тестового файла с формулами SUM, AVG и другими базовыми функциями
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
import os

def create_excel_with_formulas():
    """Создать Excel файл с различными формулами для тестирования"""
    
    # Создаем новую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Формулы"
    
    # Заголовки
    headers = ['Товар', 'Цена', 'Количество', 'Сумма', 'НДС (20%)', 'Итого']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    
    # Данные
    data = [
        ['Ноутбук', 50000, 2],
        ['Мышь', 1500, 5],
        ['Клавиатура', 3000, 3],
        ['Монитор', 25000, 1],
        ['Наушники', 8000, 2]
    ]
    
    # Заполняем данные и добавляем формулы
    for row_idx, row_data in enumerate(data, 2):
        # Основные данные
        ws.cell(row=row_idx, column=1, value=row_data[0])  # Товар
        ws.cell(row=row_idx, column=2, value=row_data[1])  # Цена
        ws.cell(row=row_idx, column=3, value=row_data[2])  # Количество
        
        # Формула: Сумма = Цена * Количество
        ws.cell(row=row_idx, column=4, value=f'=B{row_idx}*C{row_idx}')
        
        # Формула: НДС = Сумма * 0.2
        ws.cell(row=row_idx, column=5, value=f'=D{row_idx}*0.2')
        
        # Формула: Итого = Сумма + НДС
        ws.cell(row=row_idx, column=6, value=f'=D{row_idx}+E{row_idx}')
    
    # Добавляем итоговые строки с агрегатными функциями
    total_row = len(data) + 3
    
    # Заголовки для итогов
    ws.cell(row=total_row, column=1, value='ИТОГИ:')
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    
    # SUM формулы
    ws.cell(row=total_row, column=2, value=f'=SUM(B2:B{len(data)+1})')  # Общая стоимость
    ws.cell(row=total_row, column=3, value=f'=SUM(C2:C{len(data)+1})')  # Общее количество
    ws.cell(row=total_row, column=4, value=f'=SUM(D2:D{len(data)+1})')  # Общая сумма
    ws.cell(row=total_row, column=5, value=f'=SUM(E2:E{len(data)+1})')  # Общий НДС
    ws.cell(row=total_row, column=6, value=f'=SUM(F2:F{len(data)+1})')  # Общий итог
    
    # AVG формулы
    avg_row = total_row + 1
    ws.cell(row=avg_row, column=1, value='СРЕДНИЕ:')
    ws.cell(row=avg_row, column=1).font = Font(bold=True)
    
    ws.cell(row=avg_row, column=2, value=f'=AVERAGE(B2:B{len(data)+1})')  # Средняя цена
    ws.cell(row=avg_row, column=3, value=f'=AVERAGE(C2:C{len(data)+1})')  # Среднее количество
    ws.cell(row=avg_row, column=4, value=f'=AVERAGE(D2:D{len(data)+1})')  # Средняя сумма
    ws.cell(row=avg_row, column=5, value=f'=AVERAGE(E2:E{len(data)+1})')  # Средний НДС
    ws.cell(row=avg_row, column=6, value=f'=AVERAGE(F2:F{len(data)+1})')  # Средний итог
    
    # COUNT и другие функции
    count_row = avg_row + 1
    ws.cell(row=count_row, column=1, value='КОЛИЧЕСТВО:')
    ws.cell(row=count_row, column=1).font = Font(bold=True)
    
    ws.cell(row=count_row, column=2, value=f'=COUNT(B2:B{len(data)+1})')  # Количество записей
    ws.cell(row=count_row, column=3, value=f'=MAX(C2:C{len(data)+1})')   # Максимальное количество
    ws.cell(row=count_row, column=4, value=f'=MIN(D2:D{len(data)+1})')   # Минимальная сумма
    ws.cell(row=count_row, column=5, value=f'=MAX(E2:E{len(data)+1})')   # Максимальный НДС
    ws.cell(row=count_row, column=6, value=f'=MIN(F2:F{len(data)+1})')   # Минимальный итог
    
    # Условные формулы
    if_row = count_row + 2
    ws.cell(row=if_row, column=1, value='УСЛОВИЯ:')
    ws.cell(row=if_row, column=1).font = Font(bold=True)
    
    # IF формула: если цена > 10000, то "Дорого", иначе "Доступно"
    ws.cell(row=if_row, column=2, value='=IF(B2>10000,"Дорого","Доступно")')
    
    # COUNTIF: количество товаров дороже 5000
    ws.cell(row=if_row, column=3, value=f'=COUNTIF(B2:B{len(data)+1},">5000")')
    
    # SUMIF: сумма для товаров дороже 5000
    ws.cell(row=if_row, column=4, value=f'=SUMIF(B2:B{len(data)+1},">5000",D2:D{len(data)+1})')
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем файл
    filename = 'test_formulas.xlsx'
    wb.save(filename)
    print(f"✅ Создан файл с формулами: {filename}")
    
    return filename

def test_formula_loading():
    """Тестирование загрузки формул"""
    print("\n🧪 Тестирование загрузки Excel файла с формулами...")
    
    # Создаем тестовый файл
    filename = create_excel_with_formulas()
    
    if not os.path.exists(filename):
        print("❌ Файл не создан")
        return
    
    # Тестируем загрузку с data_only=True (вычисленные значения)
    print("\n📊 Загрузка с data_only=True (вычисленные значения):")
    wb_values = openpyxl.load_workbook(filename, data_only=True)
    ws_values = wb_values.active
    
    for row in ws_values.iter_rows(min_row=1, max_row=10, values_only=True):
        print(f"   {row}")
    
    # Тестируем загрузку с data_only=False (формулы как текст)
    print("\n📝 Загрузка с data_only=False (формулы как текст):")
    wb_formulas = openpyxl.load_workbook(filename, data_only=False)
    ws_formulas = wb_formulas.active
    
    for row_idx in range(1, 11):
        row_data = []
        for col_idx in range(1, 7):
            cell = ws_formulas.cell(row=row_idx, column=col_idx)
            if cell.value and str(cell.value).startswith('='):
                row_data.append(f"FORMULA: {cell.value}")
            else:
                row_data.append(cell.value)
        print(f"   {tuple(row_data)}")
    
    print(f"\n📁 Файл сохранен как: {os.path.abspath(filename)}")

if __name__ == '__main__':
    test_formula_loading()