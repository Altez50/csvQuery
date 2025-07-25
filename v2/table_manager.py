from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QListWidget,
    QListWidgetItem, QMenu, QAction, QMessageBox, QInputDialog,
    QDialog, QFormLayout, QLineEdit, QDialogButtonBox, QLabel,
    QApplication, QTableWidgetItem, QComboBox, QTreeWidget, QTreeWidgetItem,
    QFileDialog
)
from PyQt5.QtCore import Qt, pyqtSignal, QEvent
from PyQt5.QtGui import QIcon, QDrag, QPixmap, QPainter, QKeySequence
import sqlite3
import re
import os

def clean_header(header):
    """Clean header for SQLite compatibility"""
    h = header.strip().lower()
    h = re.sub(r'[\s\n\r\t]+', ' ', h)
    h = re.sub(r'[^\w\s\(\)\-\.,:;\[\]]', '', h)
    h = ''.join(ch for ch in h if ch.isprintable())
    h = re.sub(r' +', ' ', h)
    return h or 'col'

def make_unique_headers(headers):
    """Make headers unique for SQLite"""
    seen = {}
    result = []
    for idx, h in enumerate(headers):
        h_clean = clean_header(h)
        if h_clean in seen:
            seen[h_clean] += 1
            new_h = f"{h_clean}_{seen[h_clean]}"
        else:
            seen[h_clean] = 0
            new_h = h_clean
        result.append(new_h)
    return result

class TableEditDialog(QDialog):
    """Dialog for editing table properties"""
    def __init__(self, table_name="", parent=None):
        super().__init__(parent)
        self.setWindowTitle("Edit Table")
        self.setModal(True)
        self.resize(300, 150)
        
        layout = QFormLayout(self)
        
        self.name_edit = QLineEdit(table_name)
        layout.addRow("Table Name:", self.name_edit)
        
        # Buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel,
            Qt.Horizontal, self
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)
        
    def get_table_name(self):
        return self.name_edit.text().strip()

class TableManager(QWidget):
    """Table manager widget for managing database tables"""
    table_selected = pyqtSignal(str)  # Emitted when a table is selected
    table_renamed = pyqtSignal(str, str)  # Emitted when a table is renamed (old_name, new_name)
    table_deleted = pyqtSignal(str)  # Emitted when a table is deleted
    
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.convert_first_row_to_headers = True  # Option to treat first row as headers
        self.double_click_mode = "CSV editor"  # Default mode for double-click
        self.init_ui()
        
        # Install event filter for clipboard paste
        self.installEventFilter(self)
        self.setFocusPolicy(Qt.StrongFocus)
        
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # Header
        header_layout = QHBoxLayout()
        header_label = QLabel("Tables")
        header_label.setStyleSheet("font-weight: bold; font-size: 12px;")
        header_layout.addWidget(header_label)
        header_layout.addStretch()
        
        # Double-click mode switcher
        self.mode_switcher = QComboBox()
        self.mode_switcher.addItems(["CSV editor", "SQL Query"])
        self.mode_switcher.setCurrentText(self.double_click_mode)
        self.mode_switcher.setMaximumWidth(100)
        self.mode_switcher.setToolTip("Choose what happens on double-click")
        self.mode_switcher.currentTextChanged.connect(self.on_mode_changed)
        header_layout.addWidget(self.mode_switcher)
        
        # Refresh button
        self.refresh_btn = QPushButton("↻")
        self.refresh_btn.setMaximumSize(25, 25)
        self.refresh_btn.setToolTip("Refresh tables")
        self.refresh_btn.clicked.connect(self.refresh_tables)
        header_layout.addWidget(self.refresh_btn)
        
        layout.addLayout(header_layout)
        
        # Table tree (changed from list to tree for grouping)
        self.table_tree = QTreeWidget()
        self.table_tree.setHeaderLabels(["Tables"])
        self.table_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_tree.customContextMenuRequested.connect(self.show_context_menu)
        self.table_tree.itemClicked.connect(self.on_table_selected)
        self.table_tree.itemDoubleClicked.connect(self.on_table_double_clicked)
        
        # Enable drag and drop
        self.table_tree.setDragDropMode(QTreeWidget.InternalMove)
        self.table_tree.setDragEnabled(True)
        self.table_tree.setAcceptDrops(True)
        self.table_tree.setDropIndicatorShown(True)
        
        layout.addWidget(self.table_tree)
        
        # Buttons with font-based icons
        button_layout = QHBoxLayout()
        
        self.add_btn = QPushButton("➕ Add")
        self.add_btn.clicked.connect(self.add_table)
        self.add_btn.setStyleSheet("text-align: left; padding: 4px;")
        button_layout.addWidget(self.add_btn)
        
        self.paste_btn = QPushButton("📋 Paste")
        self.paste_btn.clicked.connect(self.paste_tsv_from_clipboard)
        self.paste_btn.setToolTip("Create table from tab-separated values in clipboard (Ctrl+V)")
        self.paste_btn.setStyleSheet("text-align: left; padding: 4px;")
        button_layout.addWidget(self.paste_btn)
        
        self.edit_btn = QPushButton("✏️ Edit")
        self.edit_btn.clicked.connect(self.edit_table)
        self.edit_btn.setStyleSheet("text-align: left; padding: 4px;")
        button_layout.addWidget(self.edit_btn)
        
        self.delete_btn = QPushButton("🗑️ Delete")
        self.delete_btn.clicked.connect(self.delete_table)
        self.delete_btn.setStyleSheet("text-align: left; padding: 4px;")
        button_layout.addWidget(self.delete_btn)
        
        layout.addLayout(button_layout)
        
    def eventFilter(self, source, event):
        """Handle keyboard events for clipboard paste"""
        if source == self and event.type() == QEvent.KeyPress:
            if event.matches(QKeySequence.Paste):
                self.paste_tsv_from_clipboard()
                return True
        return super().eventFilter(source, event)
        
    def paste_tsv_from_clipboard(self):
        """Create table from TSV data in clipboard"""
        clipboard = QApplication.clipboard().text()
        if not clipboard:
            QMessageBox.information(self, "Info", "Clipboard is empty")
            return
            
        try:
            # Parse TSV data
            rows = clipboard.split('\n')
            # Remove empty rows
            rows = [row for row in rows if row.strip()]
            
            if not rows:
                QMessageBox.information(self, "Info", "No data found in clipboard")
                return
                
            data = [row.split('\t') for row in rows]
            
            # Handle headers
            if self.convert_first_row_to_headers and data:
                headers_row = data[0]
                valid_headers = make_unique_headers(headers_row)
                data = data[1:]  # Remove header row from data
            else:
                # Generate column names
                max_cols = max(len(row) for row in data) if data else 0
                valid_headers = [f"col{i+1}" for i in range(max_cols)]
                
            if not data:
                QMessageBox.information(self, "Info", "No data rows found after processing headers")
                return
                
            # Generate unique table name
            table_name = self.generate_unique_table_name("pasted_data")
            
            # Create table in database
            self.create_table_from_data(table_name, valid_headers, data)
            
            # Refresh table list and select new table
            self.refresh_tables()
            self.select_table(table_name)
            
            # Load data into CSV editor if available
            if hasattr(self.main_window, 'csv_editor'):
                self.main_window.csv_editor.csv_headers = valid_headers
                self.main_window.csv_editor.csv_data = data
                self.main_window.csv_editor.update_table_display()
                
            self.main_window.log_message(f"Created table '{table_name}' from clipboard with {len(data)} rows, {len(valid_headers)} columns")
            QMessageBox.information(self, "Success", f"Table '{table_name}' created successfully from clipboard data")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to create table from clipboard: {e}")
            
    def delete_excel_group(self, group_item):
        """Delete all tables in an Excel group"""
        item_data = group_item.data(0, Qt.UserRole)
        if not item_data or item_data.get('type') != 'excel_group':
            return
            
        file_name = item_data.get('file_name', '')
        table_names = item_data.get('tables', [])
        
        if not table_names:
            return
            
        # Confirm deletion
        reply = QMessageBox.question(
            self, 'Delete Excel Group',
            f'Are you sure you want to delete the entire Excel group "{file_name}"?\n\n'
            f'This will delete {len(table_names)} tables:\n' + '\n'.join(f'• {name}' for name in table_names),
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                if not hasattr(self.main_window, 'sqlite_conn') or not self.main_window.sqlite_conn:
                    QMessageBox.warning(self, "Error", "No database connection")
                    return
                    
                cursor = self.main_window.sqlite_conn.cursor()
                
                # Delete all tables in the group
                for table_name in table_names:
                    cursor.execute(f'DROP TABLE IF EXISTS "{table_name}"')
                    self.table_deleted.emit(table_name)
                    
                self.main_window.sqlite_conn.commit()
                
                # Refresh the table tree
                self.refresh_tables()
                
                self.main_window.log_message(f"Deleted Excel group '{file_name}' with {len(table_names)} tables")
                QMessageBox.information(self, "Success", f"Excel group '{file_name}' deleted successfully")
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete Excel group: {e}")
                self.main_window.log_message(f"Failed to delete Excel group '{file_name}': {e}")
            
    def select_table(self, table_name):
        """Select a table in the table tree"""
        def find_table_item(parent_item, target_name):
            """Recursively find table item by name"""
            if parent_item is None:
                # Search top-level items
                for i in range(self.table_tree.topLevelItemCount()):
                    item = self.table_tree.topLevelItem(i)
                    if item.text(0) == target_name:
                        return item
                    # Search children
                    found = find_table_item(item, target_name)
                    if found:
                        return found
            else:
                # Search children of parent_item
                for i in range(parent_item.childCount()):
                    child = parent_item.child(i)
                    if child.text(0) == target_name:
                        return child
                    # Search grandchildren
                    found = find_table_item(child, target_name)
                    if found:
                        return found
            return None
        
        item = find_table_item(None, table_name)
        if item:
            self.table_tree.setCurrentItem(item)
            # Expand parent if needed
            parent = item.parent()
            if parent:
                parent.setExpanded(True)
            
    def generate_unique_table_name(self, base_name):
        """Generate a unique table name"""
        if not hasattr(self.main_window, 'sqlite_conn') or not self.main_window.sqlite_conn:
            return base_name
            
        try:
            cursor = self.main_window.sqlite_conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            existing_tables = {row[0] for row in cursor.fetchall()}
            
            if base_name not in existing_tables:
                return base_name
                
            # Find unique name
            counter = 1
            while f"{base_name}_{counter}" in existing_tables:
                counter += 1
            return f"{base_name}_{counter}"
            
        except Exception:
            return f"{base_name}_1"
            
    def create_table_from_data(self, table_name, headers, data):
        """Create SQLite table and insert data"""
        if not hasattr(self.main_window, 'sqlite_conn') or not self.main_window.sqlite_conn:
            # Create in-memory database if none exists
            self.main_window.sqlite_conn = sqlite3.connect(":memory:")
            
        try:
            cursor = self.main_window.sqlite_conn.cursor()
            
            # Drop table if exists
            cursor.execute(f'DROP TABLE IF EXISTS "{table_name}"')
            
            # Create table
            col_defs = ', '.join([f'"{h}" TEXT' for h in headers])
            cursor.execute(f'CREATE TABLE "{table_name}" ({col_defs})')
            
            # Insert data
            if data:
                placeholders = ','.join(['?'] * len(headers))
                # Ensure all rows have the same number of columns
                safe_rows = []
                for row in data:
                    safe_row = row[:len(headers)] + [''] * (len(headers) - len(row))
                    safe_rows.append(safe_row)
                cursor.executemany(f'INSERT INTO "{table_name}" VALUES ({placeholders})', safe_rows)
                
            self.main_window.sqlite_conn.commit()
            
        except Exception as e:
            raise Exception(f"Database error: {e}")
        
    def refresh_tables(self):
        """Refresh the table tree from database with grouping"""
        self.table_tree.clear()
        
        if not hasattr(self.main_window, 'sqlite_conn') or not self.main_window.sqlite_conn:
            return
            
        try:
            cursor = self.main_window.sqlite_conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
            tables = cursor.fetchall()
            
            # Group tables by Excel file origin
            excel_groups = {}
            standalone_tables = []
            
            for table in tables:
                table_name = table[0]
                if not table_name.startswith('sqlite_'):  # Skip system tables
                    # Check if table name follows Excel pattern: filename_sheetname
                    if '_' in table_name:
                        # Try to identify Excel file groups
                        parts = table_name.split('_')
                        if len(parts) >= 2:
                            # Assume last part is sheet name, rest is file name
                            potential_file = '_'.join(parts[:-1])
                            sheet_name = parts[-1]
                            
                            # Check if there are other tables with same file prefix
                            similar_tables = [t[0] for t in tables if t[0].startswith(potential_file + '_') and not t[0].startswith('sqlite_')]
                            
                            if len(similar_tables) > 1:
                                # This is part of an Excel group
                                if potential_file not in excel_groups:
                                    excel_groups[potential_file] = []
                                excel_groups[potential_file].append(table_name)
                            else:
                                standalone_tables.append(table_name)
                        else:
                            standalone_tables.append(table_name)
                    else:
                        standalone_tables.append(table_name)
            
            # Add Excel groups to tree
            for file_name, table_names in excel_groups.items():
                if len(table_names) > 1:  # Only create group if multiple tables
                    group_item = QTreeWidgetItem([f"📊 {file_name} ({len(table_names)} sheets)"])
                    group_item.setData(0, Qt.UserRole, {'type': 'excel_group', 'file_name': file_name, 'tables': table_names})
                    group_item.setToolTip(0, f"Excel file group: {file_name}\nRight-click to delete entire group")
                    
                    for table_name in sorted(table_names):
                        # Extract sheet name for display
                        sheet_name = table_name.replace(file_name + '_', '', 1)
                        table_item = QTreeWidgetItem([f"📋 {sheet_name}"])
                        table_item.setData(0, Qt.UserRole, {'type': 'table', 'table_name': table_name, 'sheet_name': sheet_name})
                        table_item.setToolTip(0, f"Table: {table_name}")
                        group_item.addChild(table_item)
                    
                    self.table_tree.addTopLevelItem(group_item)
                    group_item.setExpanded(True)
                else:
                    # Single table, add to standalone
                    standalone_tables.extend(table_names)
            
            # Add standalone tables
            for table_name in sorted(standalone_tables):
                item = QTreeWidgetItem([f"📋 {table_name}"])
                item.setData(0, Qt.UserRole, {'type': 'table', 'table_name': table_name})
                item.setToolTip(0, f"Table: {table_name}")
                self.table_tree.addTopLevelItem(item)
                    
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to refresh tables: {e}")
            
    def show_context_menu(self, position):
        """Show context menu for table operations"""
        item = self.table_tree.itemAt(position)
        if not item:
            return
            
        menu = QMenu(self)
        item_data = item.data(0, Qt.UserRole)
        
        if item_data and item_data.get('type') == 'excel_group':
            # Excel group context menu
            save_xlsx_action = QAction("💾 Save to XLSX", self)
            save_xlsx_action.triggered.connect(lambda: self.save_group_to_xlsx(item))
            menu.addAction(save_xlsx_action)
            
            combine_xlsx_action = QAction("📊 Combine tables", self)
            combine_xlsx_action.triggered.connect(lambda: self.combine_group_tables_to_xlsx(item))
            menu.addAction(combine_xlsx_action)
            
            menu.addSeparator()
            
            delete_group_action = QAction("🗑️ Delete Entire Group", self)
            delete_group_action.triggered.connect(lambda: self.delete_excel_group(item))
            menu.addAction(delete_group_action)
            
            menu.addSeparator()
            
            expand_action = QAction("Expand All", self)
            expand_action.triggered.connect(lambda: item.setExpanded(True))
            menu.addAction(expand_action)
            
            collapse_action = QAction("Collapse", self)
            collapse_action.triggered.connect(lambda: item.setExpanded(False))
            menu.addAction(collapse_action)
            
        else:
            # Individual table context menu
            # Edit action
            edit_action = QAction("Edit", self)
            edit_action.triggered.connect(self.edit_table)
            menu.addAction(edit_action)
            
            # Rename action
            rename_action = QAction("Rename", self)
            rename_action.triggered.connect(self.rename_table)
            menu.addAction(rename_action)
            
            menu.addSeparator()
            
            # Delete action
            delete_action = QAction("Delete", self)
            delete_action.triggered.connect(self.delete_table)
            menu.addAction(delete_action)
            
            menu.addSeparator()
            
            # View data action
            view_action = QAction("View Data", self)
            view_action.triggered.connect(self.view_table_data)
            menu.addAction(view_action)
            
            # Show structure action
            structure_action = QAction("Show Structure", self)
            structure_action.triggered.connect(self.show_table_structure)
            menu.addAction(structure_action)
            
            menu.addSeparator()
            
            # Save to XLSX action
            save_xlsx_action = QAction("💾 Save to XLSX", self)
            save_xlsx_action.triggered.connect(self.save_table_to_xlsx)
            menu.addAction(save_xlsx_action)
        
        menu.exec_(self.table_tree.mapToGlobal(position))
        
    def on_table_selected(self, item):
        """Handle table selection"""
        if item:
            item_data = item.data(0, Qt.UserRole)
            if item_data and item_data.get('type') == 'table':
                table_name = item_data.get('table_name')
                if table_name:
                    self.table_selected.emit(table_name)
            elif item_data and item_data.get('type') == 'excel_group':
                # Group selected, don't emit table_selected
                pass
            else:
                # Fallback for items without data
                table_name = item.text(0).replace('📋 ', '').replace('📊 ', '')
                self.table_selected.emit(table_name)
            
    def on_mode_changed(self, mode):
        """Handle mode switcher change"""
        self.double_click_mode = mode
        # Save setting to main window settings if available
        if hasattr(self.main_window, 'settings'):
            self.main_window.settings['table_double_click_mode'] = mode
            
    def on_table_double_clicked(self, item):
        """Handle table double-click"""
        if item:
            item_data = item.data(0, Qt.UserRole)
            if item_data and item_data.get('type') == 'table':
                table_name = item_data.get('table_name')
                if table_name:
                    if self.double_click_mode == "CSV editor":
                        self.load_table_to_csv_editor(table_name)
                    else:  # SQL Query mode
                        self.view_table_data()
            elif item_data and item_data.get('type') == 'excel_group':
                # Toggle group expansion on double-click
                item.setExpanded(not item.isExpanded())
            else:
                # Fallback for items without data
                table_name = item.text(0).replace('📋 ', '').replace('📊 ', '')
                if self.double_click_mode == "CSV editor":
                    self.load_table_to_csv_editor(table_name)
                else:  # SQL Query mode
                    self.view_table_data()
                
    def load_table_to_csv_editor(self, table_name):
        """Load table data into CSV editor"""
        if not hasattr(self.main_window, 'sqlite_conn') or not self.main_window.sqlite_conn:
            QMessageBox.warning(self, "Error", "No database connection")
            return
            
        try:
            cursor = self.main_window.sqlite_conn.cursor()
            cursor.execute(f"SELECT * FROM [{table_name}]")
            rows = cursor.fetchall()
            
            # Get column names
            cursor.execute(f"PRAGMA table_info([{table_name}])")
            columns_info = cursor.fetchall()
            headers = [col[1] for col in columns_info]  # col[1] is the column name
            
            # Convert rows to list format
            data = [list(row) for row in rows]
            
            # Load into CSV editor
            if hasattr(self.main_window, 'csv_editor'):
                self.main_window.csv_editor.csv_headers = headers
                self.main_window.csv_editor.csv_data = data
                self.main_window.csv_editor.update_table_display()
                
                # Switch to CSV editor tab
                self.main_window.editor_tabs.setCurrentWidget(self.main_window.csv_editor)
                
                self.main_window.log_message(f"Loaded table '{table_name}' into CSV editor")
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to load table data: {e}")
            
    def add_table(self):
        """Add a new table"""
        dialog = TableEditDialog(parent=self)
        if dialog.exec_() == QDialog.Accepted:
            table_name = dialog.get_table_name()
            if table_name:
                self.create_table(table_name)
                
    def create_table(self, table_name):
        """Create a new table in the database"""
        if not hasattr(self.main_window, 'sqlite_conn') or not self.main_window.sqlite_conn:
            QMessageBox.warning(self, "Error", "No database connection")
            return
            
        try:
            cursor = self.main_window.sqlite_conn.cursor()
            # Create a simple table with an ID column
            cursor.execute(f"CREATE TABLE IF NOT EXISTS [{table_name}] (id INTEGER PRIMARY KEY AUTOINCREMENT)")
            self.main_window.sqlite_conn.commit()
            
            self.refresh_tables()
            self.main_window.log_message(f"Table '{table_name}' created successfully")
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to create table: {e}")
            
    def edit_table(self):
        """Edit selected table"""
        current_item = self.table_list.currentItem()
        if not current_item:
            QMessageBox.information(self, "Info", "Please select a table to edit")
            return
            
        table_name = current_item.text()
        # For now, just show table structure
        self.show_table_structure()
        
    def rename_table(self):
        """Rename selected table"""
        current_item = self.table_list.currentItem()
        if not current_item:
            QMessageBox.information(self, "Info", "Please select a table to rename")
            return
            
        old_name = current_item.text()
        new_name, ok = QInputDialog.getText(
            self, "Rename Table", f"Enter new name for table '{old_name}':", text=old_name
        )
        
        if ok and new_name and new_name != old_name:
            if not hasattr(self.main_window, 'sqlite_conn') or not self.main_window.sqlite_conn:
                QMessageBox.warning(self, "Error", "No database connection")
                return
                
            try:
                cursor = self.main_window.sqlite_conn.cursor()
                cursor.execute(f"ALTER TABLE [{old_name}] RENAME TO [{new_name}]")
                self.main_window.sqlite_conn.commit()
                
                self.refresh_tables()
                self.table_renamed.emit(old_name, new_name)
                self.main_window.log_message(f"Table renamed from '{old_name}' to '{new_name}'")
                
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to rename table: {e}")
                
    def delete_table(self):
        """Delete selected table"""
        current_item = self.table_list.currentItem()
        if not current_item:
            QMessageBox.information(self, "Info", "Please select a table to delete")
            return
            
        table_name = current_item.text()
        
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete table '{table_name}'?\n\nThis action cannot be undone.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            if not hasattr(self.main_window, 'sqlite_conn') or not self.main_window.sqlite_conn:
                QMessageBox.warning(self, "Error", "No database connection")
                return
                
            try:
                cursor = self.main_window.sqlite_conn.cursor()
                cursor.execute(f"DROP TABLE [{table_name}]")
                self.main_window.sqlite_conn.commit()
                
                self.refresh_tables()
                self.table_deleted.emit(table_name)
                self.main_window.log_message(f"Table '{table_name}' deleted successfully")
                
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to delete table: {e}")
                
    def view_table_data(self):
        """View data in selected table"""
        current_item = self.table_list.currentItem()
        if not current_item:
            QMessageBox.information(self, "Info", "Please select a table to view")
            return
            
        table_name = current_item.text()
        
        # Switch to SQL editor and set a SELECT query
        if hasattr(self.main_window, 'sql_editor'):
            query = f"SELECT * FROM [{table_name}] LIMIT 100;"
            self.main_window.sql_editor.sql_edit.setText(query)
            self.main_window.editor_tabs.setCurrentWidget(self.main_window.sql_editor)
            
    def show_table_structure(self):
        """Show structure of selected table"""
        current_item = self.table_list.currentItem()
        if not current_item:
            QMessageBox.information(self, "Info", "Please select a table")
            return
            
        table_name = current_item.text()
        
        if not hasattr(self.main_window, 'sqlite_conn') or not self.main_window.sqlite_conn:
            QMessageBox.warning(self, "Error", "No database connection")
            return
            
        try:
            cursor = self.main_window.sqlite_conn.cursor()
            cursor.execute(f"PRAGMA table_info([{table_name}])")
            columns = cursor.fetchall()
            
            structure_text = f"Table: {table_name}\n\nColumns:\n"
            for col in columns:
                cid, name, type_, notnull, default, pk = col
                structure_text += f"  {name} ({type_})"
                if pk:
                    structure_text += " PRIMARY KEY"
                if notnull:
                    structure_text += " NOT NULL"
                if default is not None:
                    structure_text += f" DEFAULT {default}"
                structure_text += "\n"
                
            # Get row count
            cursor.execute(f"SELECT COUNT(*) FROM [{table_name}]")
            count = cursor.fetchone()[0]
            structure_text += f"\nRow count: {count}"
            
            QMessageBox.information(self, f"Table Structure - {table_name}", structure_text)
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to get table structure: {e}")
            
    def get_selected_table(self):
        """Get currently selected table name"""
        current_item = self.table_list.currentItem()
        return current_item.text() if current_item else None
        
    def select_table(self, table_name):
        """Select a table by name"""
        for i in range(self.table_list.count()):
            item = self.table_list.item(i)
            if item.text() == table_name:
                self.table_list.setCurrentItem(item)
                break
                
    def save_table_to_xlsx(self):
        """Save selected table to XLSX file"""
        current_item = self.table_tree.currentItem()
        if not current_item:
            QMessageBox.information(self, "Info", "Please select a table to export")
            return
            
        item_data = current_item.data(0, Qt.UserRole)
        if not item_data or item_data.get('type') != 'table':
            QMessageBox.information(self, "Info", "Please select a table to export")
            return
            
        table_name = item_data.get('table_name')
        if not table_name:
            QMessageBox.warning(self, "Error", "Could not determine table name")
            return
            
        # Get save file path
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Table to XLSX", f"{table_name}.xlsx", "Excel files (*.xlsx);;All files (*.*)"
        )
        
        if not file_path:
            return
            
        try:
            # Check if openpyxl is available
            try:
                import openpyxl
                from openpyxl.utils.dataframe import dataframe_to_rows
                from openpyxl.styles import Font, PatternFill
            except ImportError:
                QMessageBox.warning(self, "Error", "openpyxl library is required for XLSX export. Please install it with: pip install openpyxl")
                return
                
            if not hasattr(self.main_window, 'sqlite_conn') or not self.main_window.sqlite_conn:
                QMessageBox.warning(self, "Error", "No database connection")
                return
                
            # Get table data
            cursor = self.main_window.sqlite_conn.cursor()
            cursor.execute(f"SELECT * FROM [{table_name}]")
            rows = cursor.fetchall()
            
            # Get column names
            cursor.execute(f"PRAGMA table_info([{table_name}])")
            columns_info = cursor.fetchall()
            headers = [col[1] for col in columns_info]
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = table_name[:31]  # Excel sheet name limit
            
            # Write headers with formatting
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                
            # Write data
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
                
            # Save workbook
            wb.save(file_path)
            
            QMessageBox.information(self, "Success", f"Table '{table_name}' exported to {file_path}")
            self.main_window.log_message(f"Table '{table_name}' exported to XLSX: {file_path}")
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to export table to XLSX: {e}")
            
    def save_group_to_xlsx(self, group_item):
        """Save each table in group to its own sheet in one XLSX file"""
        if not group_item:
            return
            
        item_data = group_item.data(0, Qt.UserRole)
        if not item_data or item_data.get('type') != 'excel_group':
            QMessageBox.warning(self, "Error", "Invalid group selection")
            return
            
        group_name = group_item.text(0).replace('📁 ', '')
        
        # Get save file path
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Group to XLSX", f"{group_name}.xlsx", "Excel files (*.xlsx);;All files (*.*)"
        )
        
        if not file_path:
            return
            
        try:
            # Check if openpyxl is available
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill
            except ImportError:
                QMessageBox.warning(self, "Error", "openpyxl library is required for XLSX export. Please install it with: pip install openpyxl")
                return
                
            if not hasattr(self.main_window, 'sqlite_conn') or not self.main_window.sqlite_conn:
                QMessageBox.warning(self, "Error", "No database connection")
                return
                
            # Create workbook
            wb = openpyxl.Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            cursor = self.main_window.sqlite_conn.cursor()
            
            # Process each table in the group
            for i in range(group_item.childCount()):
                table_item = group_item.child(i)
                table_data = table_item.data(0, Qt.UserRole)
                
                if table_data and table_data.get('type') == 'table':
                    table_name = table_data.get('table_name')
                    sheet_name = table_data.get('sheet_name', table_name)
                    
                    if table_name:
                        # Get table data
                        cursor.execute(f"SELECT * FROM [{table_name}]")
                        rows = cursor.fetchall()
                        
                        # Get column names
                        cursor.execute(f"PRAGMA table_info([{table_name}])")
                        columns_info = cursor.fetchall()
                        headers = [col[1] for col in columns_info]
                        
                        # Create worksheet
                        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
                        
                        # Write headers with formatting
                        header_font = Font(bold=True)
                        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                        
                        for col_idx, header in enumerate(headers, 1):
                            cell = ws.cell(row=1, column=col_idx, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            
                        # Write data
                        for row_idx, row_data in enumerate(rows, 2):
                            for col_idx, value in enumerate(row_data, 1):
                                ws.cell(row=row_idx, column=col_idx, value=value)
                                
                        # Auto-size columns
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = adjusted_width
                            
            if len(wb.worksheets) == 0:
                QMessageBox.warning(self, "Error", "No tables found in the selected group")
                return
                
            # Save workbook
            wb.save(file_path)
            
            table_count = len(wb.worksheets)
            QMessageBox.information(self, "Success", f"Group '{group_name}' with {table_count} tables exported to {file_path}")
            self.main_window.log_message(f"Group '{group_name}' with {table_count} tables exported to XLSX: {file_path}")
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to export group to XLSX: {e}")
            
    def combine_group_tables_to_xlsx(self, group_item):
        """Combine all tables in group into one sheet"""
        if not group_item:
            return
            
        item_data = group_item.data(0, Qt.UserRole)
        if not item_data or item_data.get('type') != 'excel_group':
            QMessageBox.warning(self, "Error", "Invalid group selection")
            return
            
        group_name = group_item.text(0).replace('📁 ', '')
        
        # Get save file path
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Combine Group Tables to XLSX", f"{group_name}_combined.xlsx", "Excel files (*.xlsx);;All files (*.*)"
        )
        
        if not file_path:
            return
            
        try:
            # Check if openpyxl is available
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill
            except ImportError:
                QMessageBox.warning(self, "Error", "openpyxl library is required for XLSX export. Please install it with: pip install openpyxl")
                return
                
            if not hasattr(self.main_window, 'sqlite_conn') or not self.main_window.sqlite_conn:
                QMessageBox.warning(self, "Error", "No database connection")
                return
                
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{group_name}_combined"[:31]  # Excel sheet name limit
            
            cursor = self.main_window.sqlite_conn.cursor()
            current_row = 1
            
            # Process each table in the group
            for i in range(group_item.childCount()):
                table_item = group_item.child(i)
                table_data = table_item.data(0, Qt.UserRole)
                
                if table_data and table_data.get('type') == 'table':
                    table_name = table_data.get('table_name')
                    sheet_name = table_data.get('sheet_name', table_name)
                    
                    if table_name:
                        # Add table separator if not first table
                        if current_row > 1:
                            current_row += 2  # Add some space between tables
                            
                        # Add table title
                        title_cell = ws.cell(row=current_row, column=1, value=f"Table: {sheet_name}")
                        title_cell.font = Font(bold=True, size=14)
                        current_row += 1
                        
                        # Get table data
                        cursor.execute(f"SELECT * FROM [{table_name}]")
                        rows = cursor.fetchall()
                        
                        # Get column names
                        cursor.execute(f"PRAGMA table_info([{table_name}])")
                        columns_info = cursor.fetchall()
                        headers = [col[1] for col in columns_info]
                        
                        # Write headers with formatting
                        header_font = Font(bold=True)
                        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                        
                        for col_idx, header in enumerate(headers, 1):
                            cell = ws.cell(row=current_row, column=col_idx, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            
                        current_row += 1
                        
                        # Write data
                        for row_data in rows:
                            for col_idx, value in enumerate(row_data, 1):
                                ws.cell(row=current_row, column=col_idx, value=value)
                            current_row += 1
                            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
                
            # Save workbook
            wb.save(file_path)
            
            table_count = group_item.childCount()
            QMessageBox.information(self, "Success", f"Combined {table_count} tables from group '{group_name}' into {file_path}")
            self.main_window.log_message(f"Combined {table_count} tables from group '{group_name}' into XLSX: {file_path}")
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to combine tables to XLSX: {e}")