import csv
import pandas as pd
import re
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
                             QTableWidget, QTableWidgetItem, QFileDialog, 
                             QMessageBox, QLabel, QLineEdit, QDialog, 
                             QListWidget, QCheckBox, QDialogButtonBox,
                             QToolBar, QAction, QMenu, QApplication, QInputDialog,
                             QComboBox, QListWidgetItem, QRadioButton, QButtonGroup)
from PyQt5.QtCore import Qt, QEvent, QSize, QPoint, QRect, QUrl
from PyQt5.QtGui import QIcon, QColor, QKeySequence, QFont, QBrush, QPainter, QPen, QPixmap, QCursor
from PyQt5.QtMultimedia import QSoundEffect
import sqlite3
import colorsys
from collections import defaultdict
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from formula_engine import FormulaEngine
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def clean_header(header):
    """Clean header for SQLite compatibility"""
    h = header.strip().lower()
    h = re.sub(r'[\s\n\r\t]+', ' ', h)
    h = re.sub(r'[^\w\s\(\)\-\.,:;\[\]]', '', h)
    h = ''.join(ch for ch in h if ch.isprintable())
    h = re.sub(r' +', ' ', h)
    return h or 'col'

def make_unique_headers(headers):
    """Make headers unique for SQLite"""
    seen = {}
    result = []
    for idx, h in enumerate(headers):
        h_clean = clean_header(h)
        if h_clean in seen:
            seen[h_clean] += 1
            new_h = f"{h_clean}_{seen[h_clean]}"
        else:
            seen[h_clean] = 0
            new_h = h_clean
        result.append(new_h)
    return result

class AdvancedSearchDialog(QDialog):
    def __init__(self, headers, settings=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Advanced Search')
        self.resize(400, 350)
        layout = QVBoxLayout(self)
        
        # Select/Invert buttons
        btns_layout = QHBoxLayout()
        self.select_all_btn = QPushButton('Select All')
        self.select_all_btn.clicked.connect(self.select_all)
        self.invert_btn = QPushButton('Invert Selection')
        self.invert_btn.clicked.connect(self.invert_selection)
        btns_layout.addWidget(self.select_all_btn)
        btns_layout.addWidget(self.invert_btn)
        layout.addLayout(btns_layout)
        
        # Columns
        layout.addWidget(QLabel('Search in columns:'))
        self.columns_list = QListWidget()
        self.columns_list.setSelectionMode(QListWidget.MultiSelection)
        for col in headers:
            item = QListWidgetItem(col)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked if (not settings or col in settings.get('columns', headers)) else Qt.Unchecked)
            self.columns_list.addItem(item)
        layout.addWidget(self.columns_list)
        
        # Search mode
        layout.addWidget(QLabel('Search mode:'))
        self.mode_group = QButtonGroup(self)
        self.mode_any = QRadioButton('Any word')
        self.mode_all = QRadioButton('All words')
        self.mode_exact = QRadioButton('Exact match')
        self.mode_regex = QRadioButton('Regular expression')
        self.mode_group.addButton(self.mode_any, 0)
        self.mode_group.addButton(self.mode_all, 1)
        self.mode_group.addButton(self.mode_exact, 2)
        self.mode_group.addButton(self.mode_regex, 3)
        layout.addWidget(self.mode_any)
        layout.addWidget(self.mode_all)
        layout.addWidget(self.mode_exact)
        layout.addWidget(self.mode_regex)
        
        mode = (settings or {}).get('mode', 'any')
        if mode == 'all': self.mode_all.setChecked(True)
        elif mode == 'exact': self.mode_exact.setChecked(True)
        elif mode == 'regex': self.mode_regex.setChecked(True)
        else: self.mode_any.setChecked(True)
        
        # Case sensitivity
        self.case_checkbox = QCheckBox('Case sensitive')
        self.case_checkbox.setChecked((settings or {}).get('case', False))
        layout.addWidget(self.case_checkbox)
        
        # Buttons
        self.button_box = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        layout.addWidget(self.button_box)
        
    def get_settings(self):
        columns = [self.columns_list.item(i).text() for i in range(self.columns_list.count()) 
                  if self.columns_list.item(i).checkState() == Qt.Checked]
        mode = 'any'
        if self.mode_all.isChecked(): mode = 'all'
        elif self.mode_exact.isChecked(): mode = 'exact'
        elif self.mode_regex.isChecked(): mode = 'regex'
        return {
            'columns': columns,
            'mode': mode,
            'case': self.case_checkbox.isChecked()
        }
        
    def select_all(self):
        for i in range(self.columns_list.count()):
            self.columns_list.item(i).setCheckState(Qt.Checked)
            
    def invert_selection(self):
        for i in range(self.columns_list.count()):
            item = self.columns_list.item(i)
            item.setCheckState(Qt.Unchecked if item.checkState() == Qt.Checked else Qt.Checked)

class DragDropTableWidget(QTableWidget):
    """Custom QTableWidget with right-click region drag-and-drop functionality"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.is_drag_mode = False
        self.drag_start_pos = None
        self.drag_object = None  # Store dragged data
        self.drag_source_range = None  # Source selection range
        self.undo_data = None  # For Ctrl-Z functionality
        self.original_cursor = self.cursor()
        
        # Load sound effects
        self.take_sound = QSoundEffect()
        self.take_sound.setSource(QUrl.fromLocalFile("./sounds/take_coins.wav"))
        
        self.put_sound = QSoundEffect()
        self.put_sound.setSource(QUrl.fromLocalFile("./sounds/put_coins.wav"))
        
        # Load custom cursor
        try:
            coins_pixmap = QPixmap("./icons/coins.png")
            if not coins_pixmap.isNull():
                self.coins_cursor = QCursor(coins_pixmap)
            else:
                self.coins_cursor = QCursor(Qt.ClosedHandCursor)
        except:
            self.coins_cursor = QCursor(Qt.ClosedHandCursor)
        
    def mousePressEvent(self, event):
        """Handle mouse press events"""
        if event.button() == Qt.RightButton:
            # print("DEBUG: Right mouse button pressed!")
            # Check Scroll Lock status using QApplication.queryKeyboardModifiers()
            try:
                import win32api
                scroll_lock_on = win32api.GetKeyState(0x91) & 1  # VK_SCROLL = 0x91
                # print(f"DEBUG: mousePressEvent - Scroll Lock state: {scroll_lock_on}")
            except ImportError as e:
                # Fallback: assume Scroll Lock is always ON for testing
                scroll_lock_on = True
                # print(f"DEBUG: mousePressEvent - win32api not available ({e}), assuming Scroll Lock ON")
            
            if scroll_lock_on:
                # Enter drag mode when Scroll Lock is ON
                # print("DEBUG: Entering drag mode (Scroll Lock ON)")
                self.enter_drag_mode(event.pos())
            else:
                # Standard widget behavior when Scroll Lock is OFF
                # print("DEBUG: Using standard mouse behavior (Scroll Lock OFF)")
                super().mousePressEvent(event)
        elif event.button() == Qt.LeftButton and self.is_drag_mode:
            # Exit drag mode and perform drop
            self.exit_drag_mode(event.pos())
        else:
            super().mousePressEvent(event)
            
    def mouseMoveEvent(self, event):
        if self.is_drag_mode:
            # Handle drag selection if right button is still pressed
            if event.buttons() & Qt.RightButton and self.drag_start_pos:
                # Create selection rectangle from start to current position
                drag_rect = QRect(self.drag_start_pos, event.pos()).normalized()
                self.select_cells_in_rect(drag_rect)
                self.update_drag_object()
        else:
            super().mouseMoveEvent(event)
            
    def mouseReleaseEvent(self, event):
        if event.button() == Qt.RightButton and self.is_drag_mode:
            # Right button released - finalize drag object selection
            self.update_drag_object()
        else:
            super().mouseReleaseEvent(event)
            
    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape and self.is_drag_mode:
            # Cancel drag mode
            self.cancel_drag_mode()
        elif event.matches(QKeySequence.Undo):  # Ctrl+Z
            self.undo_last_operation()
        else:
            super().keyPressEvent(event)
            
    def contextMenuEvent(self, event):
        """Handle context menu events - suppress when Scroll Lock is ON"""
        # print("DEBUG: contextMenuEvent called!")
        # Check Scroll Lock status
        try:
            import win32api
            scroll_lock_on = win32api.GetKeyState(0x91) & 1  # VK_SCROLL = 0x91
            # print(f"DEBUG: Scroll Lock state: {scroll_lock_on}")
        except ImportError as e:
            # Fallback: assume Scroll Lock is always ON for testing
            scroll_lock_on = True
            # print(f"DEBUG: win32api not available ({e}), assuming Scroll Lock ON")
        
        if not scroll_lock_on:
            # Show context menu only when Scroll Lock is OFF
            # print("DEBUG: Showing context menu (Scroll Lock OFF)")
            super().contextMenuEvent(event)
        else:
            # When Scroll Lock is ON, do nothing (suppress context menu)
            # print("DEBUG: Suppressing context menu (Scroll Lock ON)")
            event.ignore()
            
    def select_cells_in_rect(self, rect):
        """Select all cells that intersect with the given rectangle"""
        if not rect or rect.isEmpty():
            return
            
        # Clear current selection
        self.clearSelection()
        
        # Find the range of rows and columns that intersect with the rectangle
        top_row = self.rowAt(rect.top())
        bottom_row = self.rowAt(rect.bottom())
        left_col = self.columnAt(rect.left())
        right_col = self.columnAt(rect.right())
        
        # Handle edge cases
        if top_row == -1:
            top_row = 0
        if bottom_row == -1:
            bottom_row = self.rowCount() - 1
        if left_col == -1:
            left_col = 0
        if right_col == -1:
            right_col = self.columnCount() - 1
            
        # Select the range
        if (top_row >= 0 and bottom_row >= 0 and 
            left_col >= 0 and right_col >= 0):
            for row in range(top_row, bottom_row + 1):
                for col in range(left_col, right_col + 1):
                    if row < self.rowCount() and col < self.columnCount():
                        item = self.item(row, col)
                        if not item:
                            # Create item if it doesn't exist
                            item = QTableWidgetItem("")
                            self.setItem(row, col, item)
                        item.setSelected(True)
                            
    def paintEvent(self, event):
        """Custom paint event"""
        super().paintEvent(event)
            
    def enter_drag_mode(self, pos):
        """Enter drag mode on right-click"""
        self.is_drag_mode = True
        self.drag_start_pos = pos
        
        # Play take sound
        self.take_sound.play()
        
        # Change cursor
        self.setCursor(self.coins_cursor)
        
        # Determine drag object
        clicked_row = self.rowAt(pos.y())
        clicked_col = self.columnAt(pos.x())
        
        if clicked_row >= 0 and clicked_col >= 0:
            # Check if clicked cell is in current selection
            selected_ranges = self.selectedRanges()
            cell_in_selection = False
            
            for sel_range in selected_ranges:
                if (sel_range.topRow() <= clicked_row <= sel_range.bottomRow() and
                    sel_range.leftColumn() <= clicked_col <= sel_range.rightColumn()):
                    cell_in_selection = True
                    break
            
            if not cell_in_selection:
                # Select single cell
                self.clearSelection()
                self.setCurrentCell(clicked_row, clicked_col)
                item = self.item(clicked_row, clicked_col)
                if item:
                    item.setSelected(True)
            else:
                # Cell is already in selection, update drag object immediately
                self.update_drag_object()
            
    def exit_drag_mode(self, pos):
        """Exit drag mode on left-click and perform drop"""
        if not self.is_drag_mode or not self.drag_object:
            return
            
        # Play put sound
        self.put_sound.play()
        
        # Restore cursor
        self.setCursor(self.original_cursor)
        
        # Get target position
        target_row = self.rowAt(pos.y())
        target_col = self.columnAt(pos.x())
        
        if target_row >= 0 and target_col >= 0:
            # Save current target area for undo
            self.save_undo_data(target_row, target_col)
            
            # Check if Ctrl is pressed for copy vs move
            modifiers = QApplication.keyboardModifiers()
            is_copy = modifiers & Qt.ControlModifier
            
            # Perform drop operation
            self.drop_data_at_position(target_row, target_col, is_copy)
        
        self.is_drag_mode = False
        self.drag_object = None
        self.drag_source_range = None
        
    def cancel_drag_mode(self):
        """Cancel drag mode (Escape key)"""
        self.is_drag_mode = False
        self.drag_object = None
        self.drag_source_range = None
        self.setCursor(self.original_cursor)
        
    def update_drag_object(self):
        """Update drag object from current selection"""
        selected_ranges = self.selectedRanges()
        if not selected_ranges:
            return
            
        # Find the bounding box of all selected ranges
        min_row = min(r.topRow() for r in selected_ranges)
        max_row = max(r.bottomRow() for r in selected_ranges)
        min_col = min(r.leftColumn() for r in selected_ranges)
        max_col = max(r.rightColumn() for r in selected_ranges)
        
        self.drag_source_range = {
            'top_row': min_row,
            'bottom_row': max_row,
            'left_col': min_col,
            'right_col': max_col
        }
        
        # Copy data from selected cells
        self.drag_object = []
        for row in range(min_row, max_row + 1):
            row_data = []
            for col in range(min_col, max_col + 1):
                item = self.item(row, col)
                cell_text = item.text() if item else ""
                row_data.append(cell_text)
            self.drag_object.append(row_data)
            
    def save_undo_data(self, target_row, target_col):
        """Save current state for undo functionality"""
        if not self.drag_object or not self.drag_source_range:
            return
            
        # Save source data
        source_data = []
        for row in range(self.drag_source_range['top_row'], self.drag_source_range['bottom_row'] + 1):
            row_data = []
            for col in range(self.drag_source_range['left_col'], self.drag_source_range['right_col'] + 1):
                item = self.item(row, col)
                cell_text = item.text() if item else ""
                row_data.append(cell_text)
            source_data.append(row_data)
        
        # Save target data
        target_data = []
        rows_count = len(self.drag_object)
        cols_count = len(self.drag_object[0]) if self.drag_object else 0
        
        for row_offset in range(rows_count):
            row_data = []
            for col_offset in range(cols_count):
                check_row = target_row + row_offset
                check_col = target_col + col_offset
                if check_row < self.rowCount() and check_col < self.columnCount():
                    item = self.item(check_row, check_col)
                    cell_text = item.text() if item else ""
                    row_data.append(cell_text)
                else:
                    row_data.append("")
            target_data.append(row_data)
        
        self.undo_data = {
            'source_range': self.drag_source_range.copy(),
            'source_data': source_data,
            'target_row': target_row,
            'target_col': target_col,
            'target_data': target_data
        }
        
    def drop_data_at_position(self, target_row, target_col, is_copy):
        """Drop data at target position"""
        if not self.drag_object:
            return
            
        # Paste the data
        for row_offset, row_data in enumerate(self.drag_object):
            for col_offset, cell_value in enumerate(row_data):
                paste_row = target_row + row_offset
                paste_col = target_col + col_offset
                
                # Check bounds
                if paste_row < self.rowCount() and paste_col < self.columnCount():
                    # Create item if it doesn't exist
                    item = self.item(paste_row, paste_col)
                    if not item:
                        item = QTableWidgetItem()
                        self.setItem(paste_row, paste_col, item)
                    
                    item.setText(cell_value)
        
        # If it's a move operation (not copy), clear the source cells
        if not is_copy and self.drag_source_range:
            for row in range(self.drag_source_range['top_row'], self.drag_source_range['bottom_row'] + 1):
                for col in range(self.drag_source_range['left_col'], self.drag_source_range['right_col'] + 1):
                    item = self.item(row, col)
                    if item:
                        item.setText("")
        
        # Select the dropped area
        self.clearSelection()
        rows_count = len(self.drag_object)
        cols_count = len(self.drag_object[0]) if self.drag_object else 0
        
        for row_offset in range(rows_count):
            for col_offset in range(cols_count):
                select_row = target_row + row_offset
                select_col = target_col + col_offset
                if select_row < self.rowCount() and select_col < self.columnCount():
                    item = self.item(select_row, select_col)
                    if item:
                        item.setSelected(True)
                        
    def undo_last_operation(self):
        """Undo last drag-drop operation (Ctrl+Z)"""
        if not self.undo_data:
            return
            
        # Restore source data
        source_range = self.undo_data['source_range']
        source_data = self.undo_data['source_data']
        
        for row_offset, row_data in enumerate(source_data):
            for col_offset, cell_value in enumerate(row_data):
                restore_row = source_range['top_row'] + row_offset
                restore_col = source_range['left_col'] + col_offset
                
                if restore_row < self.rowCount() and restore_col < self.columnCount():
                    item = self.item(restore_row, restore_col)
                    if not item:
                        item = QTableWidgetItem()
                        self.setItem(restore_row, restore_col, item)
                    item.setText(cell_value)
        
        # Restore target data
        target_row = self.undo_data['target_row']
        target_col = self.undo_data['target_col']
        target_data = self.undo_data['target_data']
        
        for row_offset, row_data in enumerate(target_data):
            for col_offset, cell_value in enumerate(row_data):
                restore_row = target_row + row_offset
                restore_col = target_col + col_offset
                
                if restore_row < self.rowCount() and restore_col < self.columnCount():
                    item = self.item(restore_row, restore_col)
                    if not item:
                        item = QTableWidgetItem()
                        self.setItem(restore_row, restore_col, item)
                    item.setText(cell_value)
        
        # Clear undo data
        self.undo_data = None

class CSVEditor(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.csv_data = []
        self.csv_headers = []
        self.filtered_indices = []
        self.visible_columns = []
        self.advanced_search_settings = None
        self.active_filters = {}  # Store active filters by column
        self.cell_formatting = {}  # Store cell formatting: {(row, col): {'bg_color': QColor, 'text_color': QColor, 'font': QFont}}
        self.cell_formulas = {}  # Store formula information for cells: {(row, col): formula_string}
        self.column_widths = {}  # Store Excel column width information: {col_idx: width_in_pixels}
        self.current_file = None  # Store current file path
        self.current_table_name = None  # Store current table name when loaded from database
        
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Toolbar
        toolbar = QToolBar()
        toolbar.setIconSize(QSize(20, 20))
        
        # File operations
        load_action = toolbar.addAction("📂 Load CSV")
        load_action.triggered.connect(self.load_csv)
        
        load_excel_action = toolbar.addAction("📊 Load Excel")
        load_excel_action.triggered.connect(self.load_excel_with_formatting)
        
        save_action = toolbar.addAction("💾 Save CSV")
        save_action.triggered.connect(self.save_csv)
        
        save_xlsx_action = toolbar.addAction("📊 Save XLSX")
        save_xlsx_action.triggered.connect(self.save_xlsx)
        
        toolbar.addSeparator()
        
        # Table operations
        add_row_action = toolbar.addAction("➕ Add Row")
        add_row_action.triggered.connect(self.add_row)
        
        delete_row_action = toolbar.addAction("🗑️ Delete Row")
        delete_row_action.triggered.connect(self.delete_row)
        
        toolbar.addSeparator()
        
        # Data operations
        import_to_db_action = toolbar.addAction("📊 Import to Database")
        import_to_db_action.triggered.connect(self.import_to_database)
        
        layout.addWidget(toolbar)
        
        # Search bar
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("Search:"))
        self.search_input = QLineEdit()
        self.search_input.textChanged.connect(self.perform_search)
        search_layout.addWidget(self.search_input)
        
        self.advanced_search_btn = QPushButton("Advanced")
        self.advanced_search_btn.clicked.connect(self.open_advanced_search)
        search_layout.addWidget(self.advanced_search_btn)
        
        self.clear_search_btn = QPushButton("Clear")
        self.clear_search_btn.clicked.connect(self.clear_search)
        search_layout.addWidget(self.clear_search_btn)
        
        layout.addLayout(search_layout)
        
        # Active filters display
        self.filters_layout = QHBoxLayout()
        self.filters_widget = QWidget()
        self.filters_widget.setLayout(self.filters_layout)
        self.filters_widget.setVisible(False)  # Hidden by default
        layout.addWidget(self.filters_widget)
        
        # Table widget with drag-drop functionality
        self.table = DragDropTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectItems)  # Default to column selection
        self.table.setSortingEnabled(True)
        
        # Connect header click for sorting
        self.table.horizontalHeader().sectionClicked.connect(self.sort_by_column)
        
        # Enable context menu and keyboard shortcuts
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        self.table.installEventFilter(self)
        
        # Connect signals for formula editing
        self.table.itemChanged.connect(self.on_item_changed)
        self.table.itemDoubleClicked.connect(self.on_item_double_clicked)
        
        layout.addWidget(self.table)
        
        # Status info
        self.status_label = QLabel("No data loaded")
        layout.addWidget(self.status_label)
        
    def load_csv(self):
        """Load CSV file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Load CSV File", "", "CSV files (*.csv);;All files (*.*)"
        )
        if file_path:
            self.load_csv_file(file_path)
            
    def load_excel_with_formatting(self):
        """Load Excel file with formatting options"""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "Error", 
                "openpyxl library is required to load Excel files with formatting. "
                "Please install it using: pip install openpyxl")
            return
            
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Load Excel File", "", "Excel files (*.xlsx *.xls);;All files (*.*)"
        )
        if file_path:
            # Check if we should prompt for options
            prompt_for_options = True
            if hasattr(self.main_window, 'settings'):
                prompt_for_options = self.main_window.settings.get('excel_prompt_for_options', True)
            
            if prompt_for_options:
                # Show formatting options dialog
                from excel_format_dialog import ExcelFormatDialog
                dialog = ExcelFormatDialog(self.main_window)
                if dialog.exec_() == QDialog.Accepted:
                    options = dialog.get_options()
                    self.load_excel_file_with_formatting(file_path, options)
                else:
                    # Load without formatting
                    self.load_excel_file_simple(file_path)
            else:
                # Use default settings from options
                options = self.get_default_excel_options()
                self.load_excel_file_with_formatting(file_path, options)
    
    def get_default_excel_options(self):
        """Get default Excel import options from settings"""
        if hasattr(self.main_window, 'settings'):
            settings = self.main_window.settings
            return {
                'apply_background_colors': settings.get('excel_default_apply_background_colors', True),
                'apply_text_colors': settings.get('excel_default_apply_text_colors', True),
                'apply_borders': settings.get('excel_default_apply_borders', False),
                'apply_font_family': settings.get('excel_default_apply_font_family', True),
                'apply_font_size': settings.get('excel_default_apply_font_size', True),
                'apply_font_style': settings.get('excel_default_apply_font_style', True),
                'preserve_formulas': settings.get('excel_default_preserve_formulas', False),
                'convert_dates': settings.get('excel_default_convert_dates', True)
            }
        else:
            # Default values if no settings available
            return {
                'apply_background_colors': True,
                'apply_text_colors': True,
                'apply_borders': False,
                'apply_font_family': True,
                'apply_font_size': True,
                'apply_font_style': True,
                'preserve_formulas': False,
                'convert_dates': True
            }
            
    def load_csv_file(self, file_path):
        """Load CSV file from path"""
        try:
            # Read CSV with pandas for better handling
            df = pd.read_csv(file_path, encoding='utf-8')
            self.csv_headers = list(df.columns)
            self.csv_data = df.values.tolist()
            self.cell_formulas = {}  # Clear formulas for CSV files
            self.column_widths = {}  # Clear column widths for CSV files
            self.current_file = file_path  # Store current file path
            self.current_table_name = None  # Clear table name when loading from file
            
            self.update_table_display()
            self.update_main_window_title()
            self.main_window.log_message(f"CSV loaded: {len(self.csv_data)} rows, {len(self.csv_headers)} columns")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load CSV: {e}")
            
    def load_excel_file_simple(self, file_path):
        """Load Excel file without formatting (using pandas)"""
        try:
            df = pd.read_excel(file_path)
            self.csv_headers = list(df.columns)
            self.csv_data = df.values.tolist()
            self.cell_formatting = {}  # Clear any existing formatting
            self.cell_formulas = {}  # Clear formulas for simple Excel loading
            self.column_widths = {}  # Clear column widths for simple Excel loading
            self.current_file = file_path  # Store current file path
            self.current_table_name = None  # Clear table name when loading from file
            
            self.update_table_display()
            self.update_main_window_title()
            self.main_window.log_message(f"Excel loaded: {len(self.csv_data)} rows, {len(self.csv_headers)} columns")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load Excel: {e}")
            
    def load_excel_file_with_formatting(self, file_path, options):
        """Load Excel file with formatting options using openpyxl"""
        try:
            # Load workbook with openpyxl to access formatting
            wb = openpyxl.load_workbook(file_path, data_only=not options['preserve_formulas'])
            ws = wb.active
            
            # Get data and headers
            data = []
            headers = []
            self.cell_formatting = {}
            self.cell_formulas = {}  # Clear formulas for Excel loading
            self.column_widths = {}
            
            # Extract column widths from Excel
            for col_idx, column_dimension in enumerate(ws.column_dimensions.values()):
                if column_dimension.width:
                    # Convert Excel width units to approximate pixels
                    # Excel width unit ≈ 7 pixels per unit (approximate)
                    self.column_widths[col_idx] = int(column_dimension.width * 7)
            
            # Get headers from first row
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(cell) if cell is not None else f"Column_{i+1}" for i, cell in enumerate(first_row)]
            
            # Initialize FormulaEngine if formula evaluation is enabled
            formula_engine = None
            if options.get('evaluate_formulas', False) and not options['preserve_formulas']:
                # Prepare data for FormulaEngine
                worksheet_data = []
                for row in ws.iter_rows(min_row=1, values_only=False):
                    row_data = []
                    for cell in row:
                        if cell.value is None:
                            row_data.append("")
                        elif hasattr(cell, 'data_type') and cell.data_type == 'f':  # Formula cell
                            row_data.append(cell.value)  # Keep original formula
                        else:
                            row_data.append(cell.value)
                    worksheet_data.append(row_data)
                
                formula_engine = FormulaEngine(worksheet_data)
            
            # Get data from remaining rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=0):
                row_data = []
                for col_idx, cell in enumerate(row):
                    cell_value = cell.value
                    
                    # Handle formula preservation
                    if (options['preserve_formulas'] and hasattr(cell, 'data_type') and 
                        cell.data_type == 'f' and cell.value):
                        # Store the formula in cell_formulas
                        self.cell_formulas[(row_idx, col_idx)] = cell.value
                        # Display the formula as the cell value
                        cell_value = cell.value
                    # Handle formula evaluation
                    elif (formula_engine and hasattr(cell, 'data_type') and 
                        cell.data_type == 'f' and options.get('evaluate_formulas', False)):
                        try:
                            # Evaluate formula
                            evaluated_value = formula_engine.evaluate_formula(cell_value, f"{chr(65+col_idx)}{row_idx+2}")
                            cell_value = evaluated_value
                        except Exception as e:
                            # If formula evaluation fails, keep original value or show error
                            cell_value = f"#ERROR: {str(e)}"
                    
                    if cell_value is None:
                        row_data.append("")
                    elif options['convert_dates'] and hasattr(cell_value, 'strftime'):
                        row_data.append(cell_value.strftime('%Y-%m-%d %H:%M:%S'))
                    else:
                        row_data.append(str(cell_value))
                data.append(row_data)
            
            # Extract formatting if requested
            if any([options['apply_background_colors'], options['apply_text_colors'], 
                   options['apply_font_family'], options['apply_font_size'], options['apply_font_style'],
                   options.get('show_formula_indicators', False)]):
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=0):
                    for col_idx, cell in enumerate(row):
                        if (cell.value is not None or 
                            any([cell.fill.start_color.rgb != '00000000', 
                                cell.font.color and cell.font.color.rgb != '00000000']) or
                            (options.get('show_formula_indicators', False) and 
                             hasattr(cell, 'data_type') and cell.data_type == 'f')):
                            
                            formatting = {}
                            
                            # Add formula indicator
                            if (options.get('show_formula_indicators', False) and 
                                hasattr(cell, 'data_type') and cell.data_type == 'f'):
                                formatting['is_formula'] = True
                            
                            # Mark as formula cell if preserve_formulas is enabled
                            if (options['preserve_formulas'] and hasattr(cell, 'data_type') and 
                                cell.data_type == 'f' and cell.value):
                                formatting['is_formula'] = True
                            
                            # Background color
                            if options['apply_background_colors'] and cell.fill.start_color.rgb != '00000000':
                                rgb = str(cell.fill.start_color.rgb)
                                if len(rgb) == 8:  # ARGB format
                                    rgb = rgb[2:]  # Remove alpha channel
                                if len(rgb) == 6:
                                    r = int(rgb[0:2], 16)
                                    g = int(rgb[2:4], 16)
                                    b = int(rgb[4:6], 16)
                                    formatting['bg_color'] = QColor(r, g, b)
                            
                            # Text color
                            if options['apply_text_colors'] and cell.font.color and cell.font.color.rgb != '00000000':
                                rgb = str(cell.font.color.rgb)
                                if len(rgb) == 8:  # ARGB format
                                    rgb = rgb[2:]  # Remove alpha channel
                                if len(rgb) == 6:
                                    r = int(rgb[0:2], 16)
                                    g = int(rgb[2:4], 16)
                                    b = int(rgb[4:6], 16)
                                    formatting['text_color'] = QColor(r, g, b)
                            
                            # Font formatting
                            if any([options['apply_font_family'], options['apply_font_size'], options['apply_font_style']]):
                                font = QFont()
                                
                                if options['apply_font_family'] and cell.font.name:
                                    font.setFamily(cell.font.name)
                                
                                if options['apply_font_size'] and cell.font.size:
                                    font.setPointSize(int(cell.font.size))
                                
                                if options['apply_font_style']:
                                    if cell.font.bold:
                                        font.setBold(True)
                                    if cell.font.italic:
                                        font.setItalic(True)
                                    if cell.font.underline:
                                        font.setUnderline(True)
                                
                                formatting['font'] = font
                            
                            if formatting:
                                self.cell_formatting[(row_idx, col_idx)] = formatting
            
            self.csv_headers = headers
            self.csv_data = data
            self.current_file = file_path  # Store current file path
            self.current_table_name = None  # Clear table name when loading from file
            
            self.update_table_display()
            self.update_main_window_title()
            self.main_window.log_message(f"Excel loaded with formatting: {len(self.csv_data)} rows, {len(self.csv_headers)} columns")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load Excel with formatting: {e}")
            # Fallback to simple loading
            self.load_excel_file_simple(file_path)
            
    def save_csv(self):
        """Save CSV file"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save CSV File", "", "CSV files (*.csv);;All files (*.*)"
        )
        if file_path:
            self.save_csv_file(file_path)
            
    def save_csv_file(self, file_path):
        """Save CSV file to path"""
        try:
            df = pd.DataFrame(self.csv_data, columns=self.csv_headers)
            df.to_csv(file_path, index=False, encoding='utf-8')
            self.current_file = file_path  # Update current file path
            self.update_main_window_title()
            self.main_window.log_message(f"CSV saved to {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save CSV: {e}")
            
    def save_xlsx(self):
        """Save XLSX file"""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "Error", 
                "openpyxl library is required to save Excel files. "
                "Please install it using: pip install openpyxl")
            return
            
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File", "", "Excel files (*.xlsx);;All files (*.*)"
        )
        if file_path:
            self.save_xlsx_file(file_path)
            
    def save_xlsx_file(self, file_path):
        """Save XLSX file to path with formatting preservation"""
        try:
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # Write headers
            for col_idx, header in enumerate(self.csv_headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                # Apply header formatting
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(
                    start_color='366092', end_color='366092', fill_type='solid'
                )
                cell.font = openpyxl.styles.Font(color='FFFFFF', bold=True)
            
            # Write data
            for row_idx, row_data in enumerate(self.csv_data, 2):
                for col_idx, cell_data in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_data)
                    
                    # Apply preserved formatting if available
                    formatting_key = (row_idx - 2, col_idx - 1)  # Convert to 0-based indexing
                    if formatting_key in self.cell_formatting:
                        formatting = self.cell_formatting[formatting_key]
                        
                        # Apply background color
                        if 'bg_color' in formatting:
                            qcolor = formatting['bg_color']
                            hex_color = f"{qcolor.red():02x}{qcolor.green():02x}{qcolor.blue():02x}"
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color=hex_color, end_color=hex_color, fill_type='solid'
                            )
                        
                        # Apply text color
                        if 'text_color' in formatting:
                            qcolor = formatting['text_color']
                            hex_color = f"{qcolor.red():02x}{qcolor.green():02x}{qcolor.blue():02x}"
                            cell.font = openpyxl.styles.Font(color=hex_color)
                        
                        # Apply font properties
                        if 'font' in formatting:
                            qfont = formatting['font']
                            cell.font = openpyxl.styles.Font(
                                name=qfont.family(),
                                size=qfont.pointSize(),
                                bold=qfont.bold(),
                                italic=qfont.italic()
                            )
            
            # Apply column widths if available
            for col_idx, width_pixels in self.column_widths.items():
                if col_idx < len(self.csv_headers):
                    # Convert pixels back to Excel width units (approximate)
                    excel_width = width_pixels / 7
                    column_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                    ws.column_dimensions[column_letter].width = excel_width
            
            # Auto-size columns that don't have explicit widths
            for col_idx in range(len(self.csv_headers)):
                if col_idx not in self.column_widths:
                    column_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                    # Calculate optimal width based on content
                    max_length = len(str(self.csv_headers[col_idx]))
                    for row_data in self.csv_data:
                        if col_idx < len(row_data):
                            max_length = max(max_length, len(str(row_data[col_idx])))
                    # Set width with reasonable bounds
                    optimal_width = min(max(max_length + 2, 10), 50)
                    ws.column_dimensions[column_letter].width = optimal_width
            
            # Save the workbook
            wb.save(file_path)
            self.current_file = file_path  # Update current file path
            self.update_main_window_title()
            self.main_window.log_message(f"Excel file saved to {file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save Excel file: {e}")
            
    def update_table_display(self):
        """Update table widget display"""
        if not self.csv_headers:
            self.table.clear()
            self.status_label.setText("No data loaded")
            return
            
        # Set up table
        self.table.setRowCount(len(self.csv_data))
        self.table.setColumnCount(len(self.csv_headers))
        self.table.setHorizontalHeaderLabels(self.csv_headers)
        
        # Fill data
        for row_idx, row_data in enumerate(self.csv_data):
            for col_idx, cell_data in enumerate(row_data):
                item = QTableWidgetItem(str(cell_data) if cell_data is not None else "")
                
                # Apply formatting if available
                if (row_idx, col_idx) in self.cell_formatting:
                    formatting = self.cell_formatting[(row_idx, col_idx)]
                    
                    # Apply background color
                    if 'bg_color' in formatting:
                        item.setBackground(QBrush(formatting['bg_color']))
                    
                    # Apply text color
                    if 'text_color' in formatting:
                        item.setForeground(QBrush(formatting['text_color']))
                    
                    # Apply font
                    if 'font' in formatting:
                        item.setFont(formatting['font'])
                    
                    # Apply formula indicator (green border)
                    if formatting.get('is_formula', False):
                        # Add a green background tint to indicate formula cells
                        current_bg = item.background().color() if item.background().color().isValid() else QColor(255, 255, 255)
                        formula_bg = QColor(240, 255, 240)  # Light green tint
                        # Blend the colors
                        blended = QColor(
                            int((current_bg.red() + formula_bg.red()) / 2),
                            int((current_bg.green() + formula_bg.green()) / 2),
                            int((current_bg.blue() + formula_bg.blue()) / 2)
                        )
                        item.setBackground(QBrush(blended))
                        # Add tooltip to indicate it's a formula
                        item.setToolTip("Formula cell")
                
                self.table.setItem(row_idx, col_idx, item)
        
        # Apply intelligent column width sizing
        self.apply_intelligent_column_sizing()
                
        # Update status
        self.status_label.setText(f"{len(self.csv_data)} rows, {len(self.csv_headers)} columns")
        self.main_window.row_count_label.setText(f"{len(self.csv_data)} rows")
        
    def add_row(self):
        """Add new row to table"""
        if not self.csv_headers:
            QMessageBox.warning(self, "Warning", "No CSV data loaded")
            return
            
        new_row = [""] * len(self.csv_headers)
        self.csv_data.append(new_row)
        self.update_table_display()
        
    def delete_row(self):
        """Delete selected row"""
        current_row = self.table.currentRow()
        if current_row >= 0 and current_row < len(self.csv_data):
            del self.csv_data[current_row]
            self.update_table_display()
        else:
            QMessageBox.warning(self, "Warning", "No row selected")
            
    def perform_search(self):
        """Perform search in table"""
        search_text = self.search_input.text().strip()
        if not search_text:
            self.clear_search()
            return
            
        # Use advanced search settings if available
        settings = self.advanced_search_settings or {}
        columns = settings.get('columns', self.csv_headers)
        mode = settings.get('mode', 'any')
        case_sensitive = settings.get('case', False)
        
        for row in range(self.table.rowCount()):
            row_visible = False
            
            # Get values from specified columns
            row_values = []
            for col_name in columns:
                try:
                    col_idx = self.csv_headers.index(col_name)
                    item = self.table.item(row, col_idx)
                    value = item.text() if item else ""
                    row_values.append(value)
                except ValueError:
                    continue
            
            if not row_values:
                continue
                
            # Apply search based on mode
            if mode == 'regex':
                try:
                    pattern = re.compile(search_text, 0 if case_sensitive else re.IGNORECASE)
                    row_visible = any(pattern.search(value) for value in row_values)
                except re.error:
                    # Invalid regex, fall back to simple search
                    search_lower = search_text if case_sensitive else search_text.lower()
                    row_visible = any(
                        search_lower in (value if case_sensitive else value.lower())
                        for value in row_values
                    )
            elif mode == 'exact':
                if case_sensitive:
                    row_visible = search_text in row_values
                else:
                    search_lower = search_text.lower()
                    row_visible = search_lower in [v.lower() for v in row_values]
            elif mode == 'all':
                words = search_text.split()
                if case_sensitive:
                    row_visible = all(
                        any(word in value for value in row_values)
                        for word in words
                    )
                else:
                    words_lower = [w.lower() for w in words]
                    values_lower = [v.lower() for v in row_values]
                    row_visible = all(
                        any(word in value for value in values_lower)
                        for word in words_lower
                    )
            else:  # mode == 'any' (default)
                words = search_text.split()
                if case_sensitive:
                    row_visible = any(
                        any(word in value for value in row_values)
                        for word in words
                    )
                else:
                    words_lower = [w.lower() for w in words]
                    values_lower = [v.lower() for v in row_values]
                    row_visible = any(
                        any(word in value for value in values_lower)
                        for word in words_lower
                    )
            
            self.table.setRowHidden(row, not row_visible)
            
    def open_advanced_search(self):
        """Open advanced search dialog"""
        if not self.csv_headers:
            QMessageBox.warning(self, "Warning", "No CSV data loaded")
            return
            
        dialog = AdvancedSearchDialog(self.csv_headers, self.advanced_search_settings, self)
        if dialog.exec_() == QDialog.Accepted:
            self.advanced_search_settings = dialog.get_settings()
            # Apply search with new settings
            self.perform_search()
            
    def clear_search(self):
        """Clear search and show all rows"""
        self.search_input.clear()
        for row in range(self.table.rowCount()):
            self.table.setRowHidden(row, False)
            
    def import_to_database(self):
        """Import CSV data to SQLite database"""
        if not self.csv_data or not self.csv_headers:
            QMessageBox.warning(self, "Warning", "No CSV data to import")
            return
            
        if not self.main_window.sqlite_conn:
            QMessageBox.warning(self, "Warning", "No database connection")
            return
            
        table_name, ok = QInputDialog.getText(self, "Table Name", "Enter table name:")
        if not ok or not table_name:
            return
            
        try:
            # Clean headers for SQLite
            clean_headers = make_unique_headers(self.csv_headers)
            
            # Create table
            columns_def = ", ".join([f'"{header}" TEXT' for header in clean_headers])
            create_sql = f'CREATE TABLE IF NOT EXISTS "{table_name}" ({columns_def})'
            
            cursor = self.main_window.sqlite_conn.cursor()
            cursor.execute(f'DROP TABLE IF EXISTS "{table_name}"')
            cursor.execute(create_sql)
            
            # Insert data
            placeholders = ", ".join(["?" for _ in clean_headers])
            insert_sql = f'INSERT INTO "{table_name}" VALUES ({placeholders})'
            
            for row in self.csv_data:
                cursor.execute(insert_sql, row)
                
            self.main_window.sqlite_conn.commit()
            
            # Update table manager
            if hasattr(self.main_window, 'table_manager'):
                self.main_window.table_manager.refresh_tables()
                
            self.main_window.log_message(f"Imported {len(self.csv_data)} rows to table '{table_name}'")
            QMessageBox.information(self, "Success", f"Data imported to table '{table_name}'")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to import data: {e}")
            
    def load_data(self, headers, data):
        """Load data directly into the editor"""
        self.csv_headers = headers
        self.csv_data = data
        self.cell_formatting = {}  # Clear any existing formatting
        self.cell_formulas = {}  # Clear formulas for direct data loading
        self.column_widths = {}  # Clear column widths for direct data loading
        self.current_file = None  # Clear current file since this is direct data loading
        self.current_table_name = None  # Clear current table name since this is direct data loading
        self.update_table_display()
        self.update_main_window_title()
        self.main_window.log_message(f"Data loaded: {len(self.csv_data)} rows, {len(self.csv_headers)} columns")
        
    def update_main_window_title(self):
        """Update main window title to show current file or table being edited"""
        base_title = "CSV Query Tool - VSCode Style"
        if self.current_table_name:
            # Show table name when data is loaded from database
            self.main_window.setWindowTitle(f"{base_title} - Table: {self.current_table_name}")
        elif self.current_file:
            # Show filename when data is loaded from file
            import os
            filename = os.path.basename(self.current_file)
            self.main_window.setWindowTitle(f"{base_title} - {filename}")
        else:
            self.main_window.setWindowTitle(base_title)
        
    def clear_table(self):
        """Clear table data"""
        self.csv_data = []
        self.csv_headers = []
        self.cell_formatting = {}  # Clear formatting data
        self.column_widths = {}  # Clear column width data
        self.current_file = None  # Clear current file
        self.current_table_name = None  # Clear current table name
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)
        self.status_label.setText("No data loaded")
        self.main_window.row_count_label.setText("0 rows")
        self.update_main_window_title()
        
    def apply_intelligent_column_sizing(self):
        """Apply intelligent column width sizing based on content and Excel formatting"""
        if not self.csv_headers:
            return
            
        # First, try to use Excel column widths if available
        excel_widths_applied = False
        if self.column_widths:
            for col_idx, width in self.column_widths.items():
                if col_idx < self.table.columnCount():
                    # Apply Excel width with reasonable bounds (min 50, max 400 pixels)
                    adjusted_width = max(50, min(400, width))
                    self.table.setColumnWidth(col_idx, adjusted_width)
                    excel_widths_applied = True
        
        # For columns without Excel width info, or if no Excel widths available,
        # use content-based auto-sizing
        if not excel_widths_applied:
            # Use Qt's built-in resize to contents
            self.table.resizeColumnsToContents()
            
            # Apply reasonable bounds to prevent extremely wide or narrow columns
            for col_idx in range(self.table.columnCount()):
                current_width = self.table.columnWidth(col_idx)
                
                # Calculate optimal width based on content
                header_width = self.calculate_text_width(self.csv_headers[col_idx]) + 20  # padding
                
                # Sample some data to estimate content width
                max_content_width = header_width
                sample_size = min(50, len(self.csv_data))  # Sample first 50 rows for performance
                
                for row_idx in range(sample_size):
                    if col_idx < len(self.csv_data[row_idx]):
                        cell_text = str(self.csv_data[row_idx][col_idx])
                        content_width = self.calculate_text_width(cell_text) + 10  # padding
                        max_content_width = max(max_content_width, content_width)
                
                # Apply reasonable bounds: min 80px, max 300px for auto-sized columns
                optimal_width = max(80, min(300, max_content_width))
                
                # If current width is too extreme, adjust it
                if current_width < 50 or current_width > 400:
                    self.table.setColumnWidth(col_idx, optimal_width)
        else:
            # Even with Excel widths, ensure columns without explicit widths are properly sized
            for col_idx in range(self.table.columnCount()):
                if col_idx not in self.column_widths:
                    # Auto-size this column
                    header_width = self.calculate_text_width(self.csv_headers[col_idx]) + 20
                    
                    max_content_width = header_width
                    sample_size = min(30, len(self.csv_data))
                    
                    for row_idx in range(sample_size):
                        if col_idx < len(self.csv_data[row_idx]):
                            cell_text = str(self.csv_data[row_idx][col_idx])
                            content_width = self.calculate_text_width(cell_text) + 10
                            max_content_width = max(max_content_width, content_width)
                    
                    optimal_width = max(80, min(250, max_content_width))
                    self.table.setColumnWidth(col_idx, optimal_width)
    
    def calculate_text_width(self, text):
        """Calculate approximate text width in pixels"""
        if not text:
            return 0
            
        # Use QFontMetrics to calculate text width more accurately
        font_metrics = self.table.fontMetrics()
        return font_metrics.horizontalAdvance(str(text))
         
    def import_to_database(self):
        """Import CSV data to SQLite database"""
        if not self.csv_data or not self.csv_headers:
            QMessageBox.warning(self, "Warning", "No CSV data to import")
            return
            
        if not self.main_window.sqlite_conn:
            QMessageBox.warning(self, "Warning", "No database connection")
            return
            
        table_name, ok = QInputDialog.getText(self, "Table Name", "Enter table name:")
        if not ok or not table_name:
            return
            
        try:
            # Clean headers for SQLite
            clean_headers = make_unique_headers(self.csv_headers)
            
            # Create table
            columns_def = ", ".join([f'"{header}" TEXT' for header in clean_headers])
            create_sql = f'CREATE TABLE IF NOT EXISTS "{table_name}" ({columns_def})'
            
            cursor = self.main_window.sqlite_conn.cursor()
            cursor.execute(f'DROP TABLE IF EXISTS "{table_name}"')
            cursor.execute(create_sql)
            
            # Insert data
            placeholders = ", ".join(["?" for _ in clean_headers])
            insert_sql = f'INSERT INTO "{table_name}" VALUES ({placeholders})'
            
            for row in self.csv_data:
                cursor.execute(insert_sql, row)
                
            self.main_window.sqlite_conn.commit()
            
            # Update table manager
            if hasattr(self.main_window, 'table_manager'):
                self.main_window.table_manager.refresh_tables()
                
            self.main_window.log_message(f"Imported {len(self.csv_data)} rows to table '{table_name}'")
            QMessageBox.information(self, "Success", f"Data imported to table '{table_name}'")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to import data: {e}")
            
    def eventFilter(self, source, event):
        """Handle keyboard events for Excel-like shortcuts"""
        if source == self.table and event.type() == QEvent.KeyPress:
            # Copy (Ctrl+C)
            if event.matches(QKeySequence.Copy):
                self.copy_selection()
                return True
            # Paste (Ctrl+V)
            elif event.matches(QKeySequence.Paste):
                self.paste_selection()
                return True
            # Delete selected cells (Delete key)
            elif event.key() == Qt.Key_Delete:
                self.delete_selected_cells()
                return True
            # Insert row (Insert key)
            elif event.key() == Qt.Key_Insert:
                self.add_row()
                return True
            # Delete row (Ctrl+- or Ctrl+Minus)
            elif event.modifiers() == Qt.ControlModifier and event.key() in [Qt.Key_Minus, Qt.Key_Plus]:
                if event.key() == Qt.Key_Minus:
                    self.delete_row()
                else:
                    self.add_row()
                return True
        return super().eventFilter(source, event)
        
    def show_context_menu(self, position):
        """Show context menu for table operations"""
        # print("DEBUG: show_context_menu called!")
        
        # Check Scroll Lock status - suppress menu if ON
        try:
            import win32api
            scroll_lock_on = win32api.GetKeyState(0x91) & 1  # VK_SCROLL = 0x91
            # print(f"DEBUG: show_context_menu - Scroll Lock state: {scroll_lock_on}")
        except ImportError as e:
            # Fallback: assume Scroll Lock is always ON for testing
            scroll_lock_on = True
            # print(f"DEBUG: show_context_menu - win32api not available ({e}), assuming Scroll Lock ON")
        
        if scroll_lock_on:
            # print("DEBUG: Suppressing context menu (Scroll Lock ON)")
            return  # Suppress context menu when Scroll Lock is ON
        
        # print("DEBUG: Showing context menu (Scroll Lock OFF)")
        
        if not self.csv_headers:
            return
            
        menu = QMenu(self)
        col = self.table.columnAt(position.x())
        row = self.table.rowAt(position.y())
        
        # Get column name for context-specific actions
        col_name = self.csv_headers[col] if col >= 0 and col < len(self.csv_headers) else f"col{col+1}"
        
        # Area operations (from old project)
        add_area_action = QAction("➕ Добавить область (+)", self)
        add_area_action.triggered.connect(self.add_selected_area)
        menu.addAction(add_area_action)
        
        del_area_action = QAction("➖ Удалить область (-)", self)
        del_area_action.triggered.connect(self.delete_selected_area)
        menu.addAction(del_area_action)
        
        menu.addSeparator()
        
        # Filter and group operations (from old project)
        if col >= 0:
            # Filter submenu
            filter_menu = menu.addMenu(f'🔍 Фильтр ({col_name})')
            
            # Get current cell value for context
            current_value = ""
            if row >= 0:
                item = self.table.item(row, col)
                if item:
                    current_value = item.text()
            
            if current_value:
                filter_equals = QAction(f"Равно '{current_value}'", self)
                filter_equals.triggered.connect(lambda: self.filter_by_value_context(row, col))
                filter_menu.addAction(filter_equals)
                
                filter_contains = QAction(f"Содержит '{current_value}'", self)
                filter_contains.triggered.connect(lambda: self.add_filter(col_name, current_value, "contains"))
                filter_menu.addAction(filter_contains)
                
                filter_starts = QAction(f"Начинается с '{current_value}'", self)
                filter_starts.triggered.connect(lambda: self.add_filter(col_name, current_value, "starts_with"))
                filter_menu.addAction(filter_starts)
                
                filter_ends = QAction(f"Заканчивается на '{current_value}'", self)
                filter_ends.triggered.connect(lambda: self.add_filter(col_name, current_value, "ends_with"))
                filter_menu.addAction(filter_ends)
                
                filter_not_equals = QAction(f"Не равно '{current_value}'", self)
                filter_not_equals.triggered.connect(lambda: self.add_filter(col_name, current_value, "not_equals"))
                filter_menu.addAction(filter_not_equals)
                
                filter_menu.addSeparator()
            
            # Clear filters for this column
            if col_name in self.active_filters:
                clear_column_filters = QAction(f"Очистить фильтры для {col_name}", self)
                clear_column_filters.triggered.connect(lambda: self.remove_filter(col_name))
                filter_menu.addAction(clear_column_filters)
            
            group_by_action = QAction(f"📊 Сгруппировать по значению ({col_name})", self)
            group_by_action.triggered.connect(lambda: self.group_by_value_context(col))
            menu.addAction(group_by_action)
            
            menu.addSeparator()
        
        # Coloring submenu (from old project)
        color_menu = menu.addMenu('🎨 Раскрасить')
        highlight_action = QAction('🔄 Повторяющиеся значения', self)
        highlight_action.triggered.connect(self.highlight_duplicates)
        color_menu.addAction(highlight_action)
        
        transitions_action = QAction('🔀 Переходы между значениями', self)
        transitions_action.triggered.connect(self.highlight_transitions)
        color_menu.addAction(transitions_action)
        
        menu.addSeparator()
        
        # Copy/Paste operations
        copy_action = menu.addAction("📋 Copy")
        copy_action.triggered.connect(self.copy_selection)
        
        paste_action = menu.addAction("📋 Paste")
        paste_action.triggered.connect(self.paste_selection)
        
        menu.addSeparator()
        
        # Row operations
        add_row_action = menu.addAction("➕ Add Row")
        add_row_action.triggered.connect(self.add_row)
        
        delete_row_action = menu.addAction("🗑️ Delete Row")
        delete_row_action.triggered.connect(self.delete_row)
        
        menu.addSeparator()
        
        # Column operations
        add_col_action = menu.addAction("➕ Add Column")
        add_col_action.triggered.connect(self.add_column)
        
        delete_col_action = menu.addAction("🗑️ Delete Column")
        delete_col_action.triggered.connect(self.delete_column)
        
        menu.addSeparator()
        
        # Selection mode toggle
        selection_menu = menu.addMenu('🎯 Режим выбора')
        
        current_behavior = self.table.selectionBehavior()
        
        select_items_action = QAction('Выбор ячеек', self)
        select_items_action.setCheckable(True)
        select_items_action.setChecked(current_behavior == QTableWidget.SelectItems)
        select_items_action.triggered.connect(lambda: self.table.setSelectionBehavior(QTableWidget.SelectItems))
        selection_menu.addAction(select_items_action)

        select_columns_action = QAction('Выбор столбцов', self)
        select_columns_action.setCheckable(True)
        select_columns_action.setChecked(current_behavior == QTableWidget.SelectColumns)
        select_columns_action.triggered.connect(lambda: self.table.setSelectionBehavior(QTableWidget.SelectColumns))
        selection_menu.addAction(select_columns_action)
        
        select_rows_action = QAction('Выбор строк', self)
        select_rows_action.setCheckable(True)
        select_rows_action.setChecked(current_behavior == QTableWidget.SelectRows)
        select_rows_action.triggered.connect(lambda: self.table.setSelectionBehavior(QTableWidget.SelectRows))
        selection_menu.addAction(select_rows_action)
        
  
        
        menu.addSeparator()
        
        # Clear operations
        clear_cells_action = menu.addAction("🧹 Clear Selected Cells")
        clear_cells_action.triggered.connect(self.delete_selected_cells)
        
        menu.exec_(self.table.mapToGlobal(position))
        
    def copy_selection(self):
        """Copy selected cells to clipboard"""
        selection = self.table.selectedRanges()
        if not selection:
            return
            
        # Get the first selection range
        sel_range = selection[0]
        
        # Build tab-separated text
        clipboard_text = []
        for row in range(sel_range.topRow(), sel_range.bottomRow() + 1):
            row_data = []
            for col in range(sel_range.leftColumn(), sel_range.rightColumn() + 1):
                item = self.table.item(row, col)
                cell_text = item.text() if item else ""
                row_data.append(cell_text)
            clipboard_text.append("\t".join(row_data))
            
        # Copy to clipboard
        clipboard = QApplication.clipboard()
        clipboard.setText("\n".join(clipboard_text))
        
    def paste_selection(self):
        """Paste from clipboard to selected cells"""
        clipboard = QApplication.clipboard()
        text = clipboard.text()
        
        if not text:
            return
            
        # Get current selection or cursor position
        current_row = self.table.currentRow()
        current_col = self.table.currentColumn()
        
        if current_row < 0 or current_col < 0:
            return
            
        # Parse clipboard data
        lines = text.split('\n')
        for row_offset, line in enumerate(lines):
            if not line.strip():
                continue
                
            cells = line.split('\t')
            target_row = current_row + row_offset
            
            # Expand table if needed
            while target_row >= self.table.rowCount():
                self.add_row()
                
            for col_offset, cell_value in enumerate(cells):
                target_col = current_col + col_offset
                
                # Expand columns if needed
                while target_col >= self.table.columnCount():
                    self.add_column()
                    
                # Set cell value
                item = QTableWidgetItem(str(cell_value))
                self.table.setItem(target_row, target_col, item)
                
                # Update CSV data
                if target_row < len(self.csv_data) and target_col < len(self.csv_data[target_row]):
                    self.csv_data[target_row][target_col] = str(cell_value)
                    
    def delete_selected_cells(self):
        """Clear content of selected cells"""
        selection = self.table.selectedRanges()
        if not selection:
            return
            
        for sel_range in selection:
            for row in range(sel_range.topRow(), sel_range.bottomRow() + 1):
                for col in range(sel_range.leftColumn(), sel_range.rightColumn() + 1):
                    item = QTableWidgetItem("")
                    self.table.setItem(row, col, item)
                    
                    # Update CSV data
                    if row < len(self.csv_data) and col < len(self.csv_data[row]):
                        self.csv_data[row][col] = ""
                        
    def add_column(self):
        """Add new column to table"""
        if not self.csv_headers:
            QMessageBox.warning(self, "Warning", "No CSV data loaded")
            return
            
        # Add column to headers
        new_col_name = f"Column_{len(self.csv_headers) + 1}"
        self.csv_headers.append(new_col_name)
        
        # Add column to data
        for row in self.csv_data:
            row.append("")
            
        self.update_table_display()
        
    def delete_column(self):
        """Delete selected columns"""
        selected_ranges = self.table.selectedRanges()
        if not selected_ranges:
            current_col = self.table.currentColumn()
            if current_col < 0 or current_col >= len(self.csv_headers):
                QMessageBox.warning(self, "Warning", "No column selected")
                return
            selected_cols = [current_col]
        else:
            # Get all selected columns
            selected_cols = set()
            for range_item in selected_ranges:
                for col in range(range_item.leftColumn(), range_item.rightColumn() + 1):
                    if 0 <= col < len(self.csv_headers):
                        selected_cols.add(col)
            selected_cols = sorted(selected_cols, reverse=True)  # Delete from right to left
        
        if not selected_cols:
            QMessageBox.warning(self, "Warning", "No valid columns selected")
            return
            
        # Confirm deletion
        col_names = [self.csv_headers[col] for col in selected_cols]
        reply = QMessageBox.question(
            self, "Confirm Deletion", 
            f"Delete {len(selected_cols)} column(s): {', '.join(col_names)}?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply != QMessageBox.Yes:
            return
            
        # Remove columns (from right to left to maintain indices)
        for col in selected_cols:
            # Remove from headers
            del self.csv_headers[col]
            
            # Remove from data
            for row in self.csv_data:
                if col < len(row):
                    del row[col]
                    
        self.update_table_display()
        
    def sort_by_column(self, logical_index):
        """Sort table by clicked column header"""
        if logical_index < 0 or logical_index >= len(self.csv_headers):
            return
            
        # Toggle sort order
        current_order = self.table.horizontalHeader().sortIndicatorOrder()
        new_order = Qt.DescendingOrder if current_order == Qt.AscendingOrder else Qt.AscendingOrder
        
        # Sort the table
        self.table.sortItems(logical_index, new_order)
    
    def add_selected_area(self):
        """Add area based on selection (from old project)"""
        sel = self.table.selectedRanges()
        if not sel:
            # If nothing selected but has columns, add row
            if self.table.columnCount() > 0:
                self.table.insertRow(self.table.rowCount())
            return
        rng = sel[0]
        # If width >= height, add row, else add column
        if rng.columnCount() >= rng.rowCount():
            insert_row = rng.bottomRow() + 1
            self.table.insertRow(insert_row)
        else:
            insert_col = rng.rightColumn() + 1
            self.table.insertColumn(insert_col)
            # Add header for new column
            if insert_col >= len(self.csv_headers):
                self.csv_headers.append(f"col{insert_col+1}")
            else:
                self.csv_headers.insert(insert_col, f"col{insert_col+1}")
            self.table.setHorizontalHeaderLabels(self.csv_headers)
    
    def delete_selected_area(self):
        """Delete selected area (from old project)"""
        sel = self.table.selectedRanges()
        if not sel:
            return
        rng = sel[0]
        
        # If selected area height is entire column height, delete columns
        if rng.rowCount() == self.table.rowCount():
            # Delete columns from right to left
            for col in range(rng.rightColumn(), rng.leftColumn() - 1, -1):
                self.table.removeColumn(col)
                if col < len(self.csv_headers):
                    del self.csv_headers[col]
            self.table.setHorizontalHeaderLabels(self.csv_headers)
        # If selected area width is entire row width, delete rows
        elif rng.columnCount() == self.table.columnCount():
            # Delete rows from bottom to top
            for row in range(rng.bottomRow(), rng.topRow() - 1, -1):
                self.table.removeRow(row)
        else:
            # Clear selected cells
            for row in range(rng.topRow(), rng.bottomRow() + 1):
                for col in range(rng.leftColumn(), rng.rightColumn() + 1):
                    item = self.table.item(row, col)
                    if item:
                        item.setText("")
    
    def filter_by_value_context(self, row, col):
        """Filter table by value in specific cell (from old project)"""
        if row >= 0 and col >= 0 and col < len(self.csv_headers):
            item = self.table.item(row, col)
            if item:
                filter_value = item.text()
                column_name = self.csv_headers[col]
                
                # Add filter using new system
                self.add_filter(column_name, filter_value, "equals")
    
    def group_by_value_context(self, col):
        """Group by value in column (simplified implementation)"""
        if col < 0 or col >= self.table.columnCount():
            return
        
        # Collect values and group data
        groups = defaultdict(list)
        for row in range(self.table.rowCount()):
            item = self.table.item(row, col)
            value = item.text() if item else ""
            row_data = []
            for c in range(self.table.columnCount()):
                cell_item = self.table.item(row, c)
                row_data.append(cell_item.text() if cell_item else "")
            groups[value].append(row_data)
        
        # Show simple message with group counts
        col_name = self.csv_headers[col] if col < len(self.csv_headers) else f"col{col+1}"
        group_info = "\n".join([f"{value}: {len(rows)} rows" for value, rows in groups.items()])
        QMessageBox.information(self, f"Группировка по {col_name}", group_info)
    
    def highlight_duplicates(self):
        """Highlight duplicate values (from old project)"""
        for col in range(self.table.columnCount()):
            # Collect values
            values = []
            for row in range(self.table.rowCount()):
                item = self.table.item(row, col)
                val = item.text() if item else ''
                values.append(val)
            
            # Count frequencies
            freq = defaultdict(int)
            for v in values:
                if v and v.strip():
                    freq[v] += 1
            
            # Color cache
            color_cache = {}
            prev_val = None
            
            for row, v in enumerate(values):
                item = self.table.item(row, col)
                if not item or not v or freq[v] <= 1:
                    if item:
                        item.setBackground(QColor(255, 255, 255))  # White
                    continue
                
                # Generate color by hash, but if consecutive values are same, shift hue
                if v not in color_cache:
                    base_hue = (hash(str(v)) % 360) / 360.0
                    # If consecutive, shift hue by 0.15 (54 degrees)
                    if prev_val == v:
                        hue = (base_hue + 0.15) % 1.0
                    else:
                        hue = base_hue
                    r, g, b = colorsys.hsv_to_rgb(hue, 0.7, 0.9)
                    color = QColor(int(r*255), int(g*255), int(b*255))
                    color_cache[v] = color
                else:
                    color = color_cache[v]
                
                item.setBackground(color)
                prev_val = v
    
    def highlight_transitions(self):
        """Highlight transitions between values (from old project)"""
        # Colors for transitions
        transition_colors = [
            QColor(30, 30, 120),   # dark blue
            QColor(90, 60, 30),    # dark brown
            QColor(60, 30, 90),    # dark purple
            QColor(30, 90, 60),    # dark green
            QColor(120, 60, 30),   # dark orange
            QColor(60, 60, 60),    # dark gray
        ]
        
        for col in range(self.table.columnCount()):
            prev_val = None
            color_idx = 0
            
            for row in range(self.table.rowCount()):
                item = self.table.item(row, col)
                if not item:
                    continue
                
                val = item.text()
                if row == 0:
                    prev_val = val
                    item.setForeground(QColor(0, 0, 0))  # Black
                    continue
                
                if val != prev_val:
                    color = transition_colors[color_idx % len(transition_colors)]
                    item.setForeground(color)
                    color_idx += 1
                else:
                    item.setForeground(QColor(0, 0, 0))  # Black
                
                prev_val = val
                
    def add_filter(self, column_name, filter_value, filter_type="equals"):
        """Add a filter for a specific column"""
        if column_name not in self.active_filters:
            self.active_filters[column_name] = []
        
        filter_info = {
            'value': filter_value,
            'type': filter_type
        }
        
        # Avoid duplicate filters
        if filter_info not in self.active_filters[column_name]:
            self.active_filters[column_name].append(filter_info)
            self.update_filters_display()
            self.apply_all_filters()
    
    def remove_filter(self, column_name, filter_value=None):
        """Remove filter(s) for a specific column"""
        if column_name in self.active_filters:
            if filter_value is None:
                # Remove all filters for this column
                del self.active_filters[column_name]
            else:
                # Remove specific filter
                self.active_filters[column_name] = [
                    f for f in self.active_filters[column_name] 
                    if f['value'] != filter_value
                ]
                if not self.active_filters[column_name]:
                    del self.active_filters[column_name]
            
            self.update_filters_display()
            self.apply_all_filters()
    
    def clear_all_filters(self):
        """Clear all active filters"""
        self.active_filters = {}
        self.update_filters_display()
        self.apply_all_filters()
    
    def update_filters_display(self):
        """Update the display of active filters"""
        # Clear existing filter widgets
        for i in reversed(range(self.filters_layout.count())):
            child = self.filters_layout.itemAt(i).widget()
            if child:
                child.setParent(None)
        
        if not self.active_filters:
            self.filters_widget.setVisible(False)
            return
        
        self.filters_widget.setVisible(True)
        
        # Add "Active Filters:" label
        filters_label = QLabel("Active Filters:")
        filters_label.setStyleSheet("font-weight: bold; color: #0066cc;")
        self.filters_layout.addWidget(filters_label)
        
        # Add filter chips
        for column_name, filters in self.active_filters.items():
            for filter_info in filters:
                filter_chip = self.create_filter_chip(column_name, filter_info)
                self.filters_layout.addWidget(filter_chip)
        
        # Add "Clear All" button
        clear_all_btn = QPushButton("Clear All")
        clear_all_btn.setStyleSheet(
            "QPushButton { background-color: #ff4444; color: white; "
            "border: none; padding: 4px 8px; border-radius: 3px; }"
            "QPushButton:hover { background-color: #cc3333; }"
        )
        clear_all_btn.clicked.connect(self.clear_all_filters)
        self.filters_layout.addWidget(clear_all_btn)
        
        # Add stretch to push everything to the left
        self.filters_layout.addStretch()
    
    def create_filter_chip(self, column_name, filter_info):
        """Create a filter chip widget"""
        chip_widget = QWidget()
        chip_layout = QHBoxLayout(chip_widget)
        chip_layout.setContentsMargins(0, 0, 0, 0)
        chip_layout.setSpacing(2)
        
        # Filter text
        filter_text = f"{column_name}: {filter_info['value']}"
        if filter_info['type'] != 'equals':
            filter_text = f"{column_name} {filter_info['type']} {filter_info['value']}"
        
        label = QLabel(filter_text)
        label.setStyleSheet(
            "QLabel { background-color: #e6f3ff; color: #0066cc; "
            "border: 1px solid #0066cc; padding: 2px 6px; border-radius: 3px; }"
        )
        chip_layout.addWidget(label)
        
        # Close button
        close_btn = QPushButton("×")
        close_btn.setFixedSize(16, 16)
        close_btn.setStyleSheet(
            "QPushButton { background-color: #0066cc; color: white; "
            "border: none; border-radius: 8px; font-weight: bold; font-size: 10px; }"
            "QPushButton:hover { background-color: #004499; }"
        )
        close_btn.clicked.connect(
            lambda: self.remove_filter(column_name, filter_info['value'])
        )
        chip_layout.addWidget(close_btn)
        
        return chip_widget
    
    def apply_all_filters(self):
        """Apply all active filters to the table"""
        if not self.active_filters:
            # Show all rows if no filters
            for row in range(self.table.rowCount()):
                self.table.setRowHidden(row, False)
            return
        
        # Apply filters
        for row in range(self.table.rowCount()):
            row_visible = True
            
            for column_name, filters in self.active_filters.items():
                # Find column index
                try:
                    col_idx = self.csv_headers.index(column_name)
                except ValueError:
                    continue
                
                # Check if row matches any filter for this column
                column_match = False
                item = self.table.item(row, col_idx)
                cell_value = item.text() if item else ""
                
                for filter_info in filters:
                    if self.matches_filter(cell_value, filter_info):
                        column_match = True
                        break
                
                # If no match for this column, hide the row
                if not column_match:
                    row_visible = False
                    break
            
            self.table.setRowHidden(row, not row_visible)
    
    def matches_filter(self, cell_value, filter_info):
        """Check if a cell value matches a filter"""
        filter_value = filter_info['value']
        filter_type = filter_info['type']
        
        if filter_type == 'equals':
            return cell_value == filter_value
        elif filter_type == 'contains':
            return filter_value.lower() in cell_value.lower()
        elif filter_type == 'starts_with':
            return cell_value.lower().startswith(filter_value.lower())
        elif filter_type == 'ends_with':
            return cell_value.lower().endswith(filter_value.lower())
        elif filter_type == 'not_equals':
            return cell_value != filter_value
        
        return False
    
    def on_item_changed(self, item):
        """Handle cell value changes and formula evaluation"""
        if not item:
            return
            
        row = item.row()
        col = item.column()
        new_value = item.text()
        
        # Update CSV data
        if row < len(self.csv_data) and col < len(self.csv_data[row]):
            self.csv_data[row][col] = new_value
            
        # Check if the value is a formula (starts with =)
        if new_value.startswith('='):
            try:
                # Import FormulaEngine
                from formula_engine import FormulaEngine
                
                # Create formula engine with current data
                formula_engine = FormulaEngine(self.csv_data)
                
                # Calculate cell address (A1, B2, etc.)
                cell_address = f"{chr(65 + col)}{row + 1}"
                
                # Evaluate formula
                result = formula_engine.evaluate_formula(new_value, cell_address)
                
                # Temporarily disconnect signal to avoid recursion
                self.table.itemChanged.disconnect(self.on_item_changed)
                
                # Update display with result but keep formula in data
                display_item = QTableWidgetItem(str(result))
                
                # Mark as formula cell with special formatting
                display_item.setBackground(QBrush(QColor(240, 255, 240)))  # Light green
                display_item.setToolTip(f"Formula: {new_value}\nResult: {result}")
                
                # Store original formula in cell formatting
                if not hasattr(self, 'cell_formulas'):
                    self.cell_formulas = {}
                self.cell_formulas[(row, col)] = new_value
                
                self.table.setItem(row, col, display_item)
                
                # Reconnect signal
                self.table.itemChanged.connect(self.on_item_changed)
                
                # Update other cells that might reference this cell
                self.update_dependent_formulas()
                
            except Exception as e:
                # Show error in cell
                self.table.itemChanged.disconnect(self.on_item_changed)
                error_item = QTableWidgetItem(f"#ERROR: {str(e)}")
                error_item.setBackground(QBrush(QColor(255, 240, 240)))  # Light red
                error_item.setToolTip(f"Formula: {new_value}\nError: {str(e)}")
                self.table.setItem(row, col, error_item)
                self.table.itemChanged.connect(self.on_item_changed)
        else:
            # Regular value, clear any formula formatting
            if hasattr(self, 'cell_formulas') and (row, col) in self.cell_formulas:
                del self.cell_formulas[(row, col)]
            
            # Reset background color for non-formula cells
            item.setBackground(QBrush(QColor(255, 255, 255)))
            item.setToolTip("")
    
    def update_dependent_formulas(self):
        """Update all formula cells that might be affected by data changes"""
        if not hasattr(self, 'cell_formulas'):
            return
            
        try:
            from formula_engine import FormulaEngine
            formula_engine = FormulaEngine(self.csv_data)
            
            # Temporarily disconnect signal
            self.table.itemChanged.disconnect(self.on_item_changed)
            
            # Re-evaluate all formula cells
            for (row, col), formula in self.cell_formulas.items():
                try:
                    cell_address = f"{chr(65 + col)}{row + 1}"
                    result = formula_engine.evaluate_formula(formula, cell_address)
                    
                    # Update display
                    display_item = QTableWidgetItem(str(result))
                    display_item.setBackground(QBrush(QColor(240, 255, 240)))
                    display_item.setToolTip(f"Formula: {formula}\nResult: {result}")
                    self.table.setItem(row, col, display_item)
                    
                except Exception as e:
                    # Show error
                    error_item = QTableWidgetItem(f"#ERROR: {str(e)}")
                    error_item.setBackground(QBrush(QColor(255, 240, 240)))
                    error_item.setToolTip(f"Formula: {formula}\nError: {str(e)}")
                    self.table.setItem(row, col, error_item)
            
            # Reconnect signal
            self.table.itemChanged.connect(self.on_item_changed)
            
        except Exception as e:
            # Reconnect signal even if there's an error
             self.table.itemChanged.connect(self.on_item_changed)
             print(f"Error updating dependent formulas: {e}")
    
    def on_item_double_clicked(self, item):
        """Handle double-click on formula cells to edit the formula"""
        if not item:
            return
            
        row = item.row()
        col = item.column()
        
        # Check if this cell contains a formula
        if hasattr(self, 'cell_formulas') and (row, col) in self.cell_formulas:
            formula = self.cell_formulas[(row, col)]
            
            # Temporarily disconnect signal to avoid recursion
            self.table.itemChanged.disconnect(self.on_item_changed)
            
            # Set the formula text for editing
            item.setText(formula)
            
            # Reconnect signal
            self.table.itemChanged.connect(self.on_item_changed)