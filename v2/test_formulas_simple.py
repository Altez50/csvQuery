#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Создание простого Excel файла с формулами для тестирования
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def create_test_excel_with_formulas():
    """Создает Excel файл с тестовыми данными и формулами"""
    
    # Создаем новую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Formulas"
    
    # Заголовки
    ws['A1'] = 'Число A'
    ws['B1'] = 'Число B'
    ws['C1'] = 'Сумма (A+B)'
    ws['D1'] = 'Произведение (A*B)'
    ws['E1'] = 'Максимум'
    
    # Тестовые данные
    test_data = [
        [10, 20],
        [5, 15],
        [8, 12],
        [25, 5],
        [30, 10]
    ]
    
    # Заполняем данные
    for i, (a, b) in enumerate(test_data, start=2):
        ws[f'A{i}'] = a
        ws[f'B{i}'] = b
        
        # Добавляем формулы
        ws[f'C{i}'] = f'=A{i}+B{i}'  # Сумма
        ws[f'D{i}'] = f'=A{i}*B{i}'  # Произведение
        ws[f'E{i}'] = f'=MAX(A{i},B{i})'  # Максимум
    
    # Добавляем итоговые формулы
    last_row = len(test_data) + 1
    summary_row = last_row + 2
    
    ws[f'A{summary_row}'] = 'Итого:'
    ws[f'B{summary_row}'] = f'=SUM(A2:A{last_row})'
    ws[f'C{summary_row}'] = f'=SUM(C2:C{last_row})'
    ws[f'D{summary_row}'] = f'=SUM(D2:D{last_row})'
    ws[f'E{summary_row}'] = f'=AVERAGE(E2:E{last_row})'
    
    # Добавляем условные формулы
    condition_row = summary_row + 2
    ws[f'A{condition_row}'] = 'Условие:'
    ws[f'B{condition_row}'] = 'Количество A>10:'
    ws[f'C{condition_row}'] = f'=COUNTIF(A2:A{last_row},">10")'
    
    # Сохраняем файл
    filename = 'test_formulas_simple.xlsx'
    wb.save(filename)
    print(f"Создан файл: {filename}")
    
    return filename

if __name__ == '__main__':
    create_test_excel_with_formulas()
    print("\nТестовый файл создан!")
    print("Откройте его в CSV Editor для тестирования формул.")