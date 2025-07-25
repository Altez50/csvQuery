#!/usr/bin/env python3
"""
Test script for intelligent column width sizing in CSV Query Tool.
This script creates sample Excel files with different column widths and content
to demonstrate the new column sizing feature.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
import os

def create_sample_excel_with_widths():
    """Create a sample Excel file with custom column widths and varied content"""
    
    # Sample data with varying content lengths
    data = {
        'ID': [1, 2, 3, 4, 5],
        'Name': ['John Doe', 'Jane Smith', 'Robert Johnson', 'Alice Brown', 'Michael Wilson'],
        'Email': ['john.doe@email.com', 'jane.smith@company.org', 'robert.johnson@business.net', 'alice.brown@corp.com', 'michael.wilson@enterprise.biz'],
        'Department': ['IT', 'Sales', 'Human Resources', 'Marketing', 'Finance'],
        'Salary': [75000, 65000, 55000, 60000, 80000],
        'Description': [
            'Software Engineer with 5+ years experience',
            'Sales Representative',
            'HR Manager responsible for recruitment and employee relations',
            'Marketing Specialist',
            'Senior Financial Analyst with expertise in budget planning and forecasting'
        ],
        'Code': ['A1', 'B2', 'C3', 'D4', 'E5'],
        'Status': ['Active', 'Active', 'On Leave', 'Active', 'Active']
    }
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Write to Excel file
    excel_file = 'sample_column_widths.xlsx'
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Employees', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Employees']
        
        # Set custom column widths (in Excel units)
        column_widths = {
            'A': 8,   # ID - narrow
            'B': 20,  # Name - medium
            'C': 35,  # Email - wide
            'D': 18,  # Department - medium
            'E': 12,  # Salary - narrow-medium
            'F': 50,  # Description - very wide
            'G': 10,  # Code - narrow
            'H': 15   # Status - medium
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Add some formatting to make it more interesting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # Format headers
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
        
        # Add some cell formatting
        salary_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        for row in range(2, len(df) + 2):
            worksheet.cell(row=row, column=5).fill = salary_fill  # Salary column
    
    print(f"✅ Created Excel file with custom column widths: {excel_file}")
    print(f"   📊 Column widths set:")
    for col, width in column_widths.items():
        col_name = df.columns[ord(col) - ord('A')]
        print(f"      {col} ({col_name}): {width} Excel units ≈ {width * 7} pixels")
    
    return excel_file

def create_sample_csv_for_comparison():
    """Create a CSV file with the same data for comparison"""
    
    data = {
        'ID': [1, 2, 3, 4, 5],
        'Name': ['John Doe', 'Jane Smith', 'Robert Johnson', 'Alice Brown', 'Michael Wilson'],
        'Email': ['john.doe@email.com', 'jane.smith@company.org', 'robert.johnson@business.net', 'alice.brown@corp.com', 'michael.wilson@enterprise.biz'],
        'Department': ['IT', 'Sales', 'Human Resources', 'Marketing', 'Finance'],
        'Salary': [75000, 65000, 55000, 60000, 80000],
        'Description': [
            'Software Engineer with 5+ years experience',
            'Sales Representative',
            'HR Manager responsible for recruitment and employee relations',
            'Marketing Specialist',
            'Senior Financial Analyst with expertise in budget planning and forecasting'
        ],
        'Code': ['A1', 'B2', 'C3', 'D4', 'E5'],
        'Status': ['Active', 'Active', 'On Leave', 'Active', 'Active']
    }
    
    df = pd.DataFrame(data)
    csv_file = 'sample_column_widths.csv'
    df.to_csv(csv_file, index=False)
    
    print(f"✅ Created CSV file for comparison: {csv_file}")
    return csv_file

def analyze_excel_column_widths(excel_file):
    """Analyze and display Excel column width information"""
    print(f"\n🔍 Analyzing Excel column widths in {excel_file}...")
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        print(f"   📋 Worksheet: {ws.title}")
        print(f"   📊 Column width analysis:")
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        
        for col_idx, (col_letter, column_dimension) in enumerate(ws.column_dimensions.items()):
            header_name = headers[col_idx] if col_idx < len(headers) else f"Column {col_idx + 1}"
            width = column_dimension.width if column_dimension.width else "Auto"
            
            if isinstance(width, (int, float)):
                pixels = int(width * 7)  # Convert to approximate pixels
                print(f"      {col_letter} ({header_name}): {width} Excel units → {pixels} pixels")
            else:
                print(f"      {col_letter} ({header_name}): {width} (will use content-based sizing)")
        
        # Sample some content to show what the algorithm will consider
        print(f"\n   📝 Sample content analysis:")
        for col_idx, header in enumerate(headers[:4]):  # Show first 4 columns
            print(f"      {header}:")
            for row_idx in range(2, min(4, ws.max_row + 1)):  # Show 2 sample rows
                cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
                if cell_value:
                    content_length = len(str(cell_value))
                    print(f"        Row {row_idx}: '{cell_value}' ({content_length} chars)")
        
    except Exception as e:
        print(f"❌ Error analyzing Excel file: {e}")

def main():
    """Main test function"""
    print("🧪 Testing Intelligent Column Width Sizing")
    print("=" * 50)
    
    # Create test files
    excel_file = create_sample_excel_with_widths()
    csv_file = create_sample_csv_for_comparison()
    
    # Analyze Excel column widths
    analyze_excel_column_widths(excel_file)
    
    print("\n✅ Test files created successfully!")
    print("\n📝 Manual Testing Instructions:")
    print("   1. Open the CSV Query Tool application")
    print(f"   2. Load the Excel file: {excel_file}")
    print("      - Choose 'Load Excel' and enable formatting options")
    print("      - Observe how column widths are preserved from Excel")
    print(f"   3. Load the CSV file: {csv_file}")
    print("      - Choose 'Load CSV'")
    print("      - Observe how columns are auto-sized based on content")
    print("   4. Compare the column widths between Excel and CSV loading")
    print("\n🎯 Expected Results:")
    print("   📊 Excel loading: Columns should respect original Excel widths")
    print("      - ID column: narrow (~56px)")
    print("      - Name column: medium (~140px)")
    print("      - Email column: wide (~245px)")
    print("      - Description column: very wide (~350px)")
    print("   📄 CSV loading: Columns should be auto-sized based on content")
    print("      - All columns sized optimally for their content")
    print("      - No column too narrow or excessively wide")
    print("\n🔧 Features to Test:")
    print("   ✓ Excel column width preservation")
    print("   ✓ Content-based auto-sizing for CSV")
    print("   ✓ Reasonable bounds (min 50px, max 400px)")
    print("   ✓ Mixed sizing (Excel widths + auto-sizing for missing columns)")

if __name__ == "__main__":
    main()