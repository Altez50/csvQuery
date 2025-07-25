#!/usr/bin/env python3
"""
Test script for Table Manager XLSX export functionality.
This script demonstrates the new context menu options for exporting tables to XLSX format.
"""

import sqlite3
import os
import pandas as pd

def create_test_database():
    """Create a test database with sample tables for testing XLSX export"""
    db_path = "test_table_export.db"
    
    # Remove existing database
    if os.path.exists(db_path):
        os.remove(db_path)
    
    # Create new database
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Create sample tables
    
    # Table 1: Employee data
    cursor.execute("""
        CREATE TABLE employees (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            department TEXT,
            salary REAL,
            hire_date TEXT
        )
    """)
    
    employees_data = [
        (1, 'John Smith', 'Engineering', 75000.0, '2020-01-15'),
        (2, 'Jane Doe', 'Marketing', 65000.0, '2019-03-22'),
        (3, 'Bob Johnson', 'Engineering', 80000.0, '2021-06-10'),
        (4, 'Alice Brown', 'HR', 55000.0, '2018-11-05'),
        (5, 'Charlie Wilson', 'Sales', 70000.0, '2020-09-18')
    ]
    
    cursor.executemany("INSERT INTO employees VALUES (?, ?, ?, ?, ?)", employees_data)
    
    # Table 2: Products data
    cursor.execute("""
        CREATE TABLE products (
            product_id INTEGER PRIMARY KEY,
            product_name TEXT NOT NULL,
            category TEXT,
            price REAL,
            stock_quantity INTEGER
        )
    """)
    
    products_data = [
        (1, 'Laptop Pro', 'Electronics', 1299.99, 25),
        (2, 'Wireless Mouse', 'Electronics', 29.99, 150),
        (3, 'Office Chair', 'Furniture', 199.99, 45),
        (4, 'Desk Lamp', 'Furniture', 49.99, 80),
        (5, 'Notebook Set', 'Stationery', 12.99, 200)
    ]
    
    cursor.executemany("INSERT INTO products VALUES (?, ?, ?, ?, ?)", products_data)
    
    # Table 3: Sales data
    cursor.execute("""
        CREATE TABLE sales (
            sale_id INTEGER PRIMARY KEY,
            employee_id INTEGER,
            product_id INTEGER,
            quantity INTEGER,
            sale_date TEXT,
            total_amount REAL
        )
    """)
    
    sales_data = [
        (1, 5, 1, 2, '2023-01-15', 2599.98),
        (2, 5, 2, 5, '2023-01-16', 149.95),
        (3, 2, 3, 1, '2023-01-17', 199.99),
        (4, 5, 4, 3, '2023-01-18', 149.97),
        (5, 2, 5, 10, '2023-01-19', 129.90)
    ]
    
    cursor.executemany("INSERT INTO sales VALUES (?, ?, ?, ?, ?, ?)", sales_data)
    
    # Create tables that simulate Excel import groups
    # Group 1: Financial data
    cursor.execute("""
        CREATE TABLE financial_q1_revenue (
            month TEXT,
            revenue REAL,
            expenses REAL,
            profit REAL
        )
    """)
    
    financial_q1_data = [
        ('January', 150000.0, 120000.0, 30000.0),
        ('February', 165000.0, 125000.0, 40000.0),
        ('March', 180000.0, 130000.0, 50000.0)
    ]
    
    cursor.executemany("INSERT INTO financial_q1_revenue VALUES (?, ?, ?, ?)", financial_q1_data)
    
    cursor.execute("""
        CREATE TABLE financial_q1_expenses (
            category TEXT,
            amount REAL,
            percentage REAL
        )
    """)
    
    expenses_data = [
        ('Salaries', 80000.0, 65.0),
        ('Rent', 15000.0, 12.0),
        ('Utilities', 8000.0, 6.5),
        ('Marketing', 12000.0, 10.0),
        ('Other', 8000.0, 6.5)
    ]
    
    cursor.executemany("INSERT INTO financial_q1_expenses VALUES (?, ?, ?)", expenses_data)
    
    # Group 2: Inventory data
    cursor.execute("""
        CREATE TABLE inventory_electronics (
            item_name TEXT,
            sku TEXT,
            quantity INTEGER,
            unit_cost REAL
        )
    """)
    
    electronics_data = [
        ('Smartphone X', 'SPX-001', 50, 299.99),
        ('Tablet Pro', 'TAB-002', 30, 399.99),
        ('Headphones', 'HP-003', 100, 79.99),
        ('Smartwatch', 'SW-004', 25, 199.99)
    ]
    
    cursor.executemany("INSERT INTO inventory_electronics VALUES (?, ?, ?, ?)", electronics_data)
    
    cursor.execute("""
        CREATE TABLE inventory_furniture (
            item_name TEXT,
            sku TEXT,
            quantity INTEGER,
            unit_cost REAL
        )
    """)
    
    furniture_data = [
        ('Executive Desk', 'ED-001', 15, 599.99),
        ('Ergonomic Chair', 'EC-002', 40, 299.99),
        ('Bookshelf', 'BS-003', 20, 149.99),
        ('Filing Cabinet', 'FC-004', 25, 199.99)
    ]
    
    cursor.executemany("INSERT INTO inventory_furniture VALUES (?, ?, ?, ?)", furniture_data)
    
    conn.commit()
    conn.close()
    
    print(f"Test database created: {db_path}")
    print("\nTables created:")
    print("- employees (5 records)")
    print("- products (5 records)")
    print("- sales (5 records)")
    print("- financial_q1_revenue (3 records)")
    print("- financial_q1_expenses (5 records)")
    print("- inventory_electronics (4 records)")
    print("- inventory_furniture (4 records)")
    
    return db_path

def analyze_database(db_path):
    """Analyze the test database structure"""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Get all tables
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
    tables = cursor.fetchall()
    
    print("\n=== Database Analysis ===")
    for table_name, in tables:
        cursor.execute(f"SELECT COUNT(*) FROM [{table_name}]")
        count = cursor.fetchone()[0]
        
        cursor.execute(f"PRAGMA table_info([{table_name}])")
        columns = cursor.fetchall()
        column_names = [col[1] for col in columns]
        
        print(f"\nTable: {table_name}")
        print(f"  Rows: {count}")
        print(f"  Columns: {', '.join(column_names)}")
    
    conn.close()

def create_sample_xlsx_files():
    """Create sample XLSX files to compare with exports"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("openpyxl not available, skipping sample XLSX creation")
        return
    
    # Create a sample multi-sheet workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Sheet 1: Employees
    ws1 = wb.create_sheet("Employees")
    headers1 = ['ID', 'Name', 'Department', 'Salary', 'Hire Date']
    
    # Write headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for col_idx, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Sample data
    data1 = [
        [1, 'John Smith', 'Engineering', 75000, '2020-01-15'],
        [2, 'Jane Doe', 'Marketing', 65000, '2019-03-22']
    ]
    
    for row_idx, row_data in enumerate(data1, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=value)
    
    # Sheet 2: Products
    ws2 = wb.create_sheet("Products")
    headers2 = ['Product ID', 'Product Name', 'Category', 'Price', 'Stock']
    
    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    data2 = [
        [1, 'Laptop Pro', 'Electronics', 1299.99, 25],
        [2, 'Wireless Mouse', 'Electronics', 29.99, 150]
    ]
    
    for row_idx, row_data in enumerate(data2, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-size columns for both sheets
    for ws in [ws1, ws2]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save("sample_multi_sheet_export.xlsx")
    print("\nSample XLSX file created: sample_multi_sheet_export.xlsx")

def print_testing_instructions():
    """Print manual testing instructions"""
    print("\n" + "="*60)
    print("MANUAL TESTING INSTRUCTIONS")
    print("="*60)
    
    print("\n1. SETUP:")
    print("   - Start the CSV Query Tool application")
    print("   - Load the test database: test_table_export.db")
    print("   - Navigate to the Tables tab in the left sidebar")
    
    print("\n2. TEST INDIVIDUAL TABLE EXPORT:")
    print("   - Right-click on any individual table (e.g., 'employees')")
    print("   - Select '💾 Save to XLSX' from context menu")
    print("   - Choose save location and filename")
    print("   - Verify the exported file contains:")
    print("     * Proper headers with bold formatting")
    print("     * All table data")
    print("     * Auto-sized columns")
    print("     * Single worksheet named after the table")
    
    print("\n3. TEST GROUP EXPORT (Multiple Sheets):")
    print("   - If you have Excel groups in the tree, right-click on a group")
    print("   - Select '💾 Save to XLSX' from context menu")
    print("   - Choose save location and filename")
    print("   - Verify the exported file contains:")
    print("     * Multiple worksheets (one per table in group)")
    print("     * Each sheet named after its source table")
    print("     * Proper formatting on all sheets")
    print("     * All data preserved")
    
    print("\n4. TEST COMBINED EXPORT (Single Sheet):")
    print("   - Right-click on a group with multiple tables")
    print("   - Select '📊 Combine tables' from context menu")
    print("   - Choose save location and filename")
    print("   - Verify the exported file contains:")
    print("     * Single worksheet with all tables combined")
    print("     * Table separators with table names")
    print("     * Headers for each table section")
    print("     * Proper spacing between tables")
    
    print("\n5. EXPECTED RESULTS:")
    print("   ✓ Files should open correctly in Excel/LibreOffice")
    print("   ✓ Headers should be bold with gray background")
    print("   ✓ Columns should be auto-sized for readability")
    print("   ✓ Data types should be preserved (numbers, text, dates)")
    print("   ✓ Success messages should appear after export")
    print("   ✓ Log messages should be recorded in the application")
    
    print("\n6. ERROR HANDLING:")
    print("   - Test with no table selected (should show info message)")
    print("   - Test with empty tables (should create file with headers only)")
    print("   - Test with very long table/column names (should be truncated)")
    
    print("\n7. COMPARISON:")
    print("   - Compare exported files with sample_multi_sheet_export.xlsx")
    print("   - Verify formatting consistency")
    print("   - Check data integrity")

if __name__ == "__main__":
    print("Creating test database for Table Manager XLSX export testing...")
    
    # Create test database
    db_path = create_test_database()
    
    # Analyze database
    analyze_database(db_path)
    
    # Create sample files
    create_sample_xlsx_files()
    
    # Print testing instructions
    print_testing_instructions()
    
    print("\n" + "="*60)
    print("Test setup complete! You can now test the XLSX export functionality.")
    print("="*60)