#!/usr/bin/env python3
"""
Test script to create an Excel file with formulas for testing the main application
"""

import openpyxl
import os

def create_test_excel_with_formulas():
    """Create a test Excel file with various formulas"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formula Test"
    
    # Add headers
    ws['A1'] = 'Value A'
    ws['B1'] = 'Value B'
    ws['C1'] = 'Sum (A+B)'
    ws['D1'] = 'Product (A*B)'
    ws['E1'] = 'Max(A,B)'
    ws['F1'] = 'Average'
    
    # Add sample data
    ws['A2'] = 10
    ws['B2'] = 20
    ws['A3'] = 15
    ws['B3'] = 25
    ws['A4'] = 5
    ws['B4'] = 30
    ws['A5'] = 12
    ws['B5'] = 18
    
    # Add formulas
    ws['C2'] = '=A2+B2'  # Simple addition
    ws['C3'] = '=A3+B3'
    ws['C4'] = '=A4+B4'
    ws['C5'] = '=A5+B5'
    
    ws['D2'] = '=A2*B2'  # Simple multiplication
    ws['D3'] = '=A3*B3'
    ws['D4'] = '=A4*B4'
    ws['D5'] = '=A5*B5'
    
    ws['E2'] = '=MAX(A2,B2)'  # MAX function
    ws['E3'] = '=MAX(A3,B3)'
    ws['E4'] = '=MAX(A4,B4)'
    ws['E5'] = '=MAX(A5,B5)'
    
    ws['F2'] = '=AVERAGE(A2:B2)'  # AVERAGE function
    ws['F3'] = '=AVERAGE(A3:B3)'
    ws['F4'] = '=AVERAGE(A4:B4)'
    ws['F5'] = '=AVERAGE(A5:B5)'
    
    # Add some summary formulas
    ws['A7'] = 'Total Sum:'
    ws['B7'] = '=SUM(C2:C5)'
    
    ws['A8'] = 'Total Average:'
    ws['B8'] = '=AVERAGE(F2:F5)'
    
    ws['A9'] = 'Count:'
    ws['B9'] = '=COUNT(A2:A5)'
    
    # Save the file
    filename = 'test_main_app_formulas.xlsx'
    wb.save(filename)
    print(f"Created test Excel file: {filename}")
    print("\nThis file contains:")
    print("- Simple arithmetic formulas (A+B, A*B)")
    print("- Function formulas (MAX, AVERAGE, SUM, COUNT)")
    print("- Cell range references (A2:B2, C2:C5, etc.)")
    print("\nTo test in the main application:")
    print("1. Run main.py")
    print("2. Open this Excel file")
    print("3. Choose 'Yes' for formatting options")
    print("4. Enable 'Preserve Formulas' in the dialog")
    print("5. Check that formula cells have green background")
    print("6. Double-click formula cells to see the formulas")
    print("7. Try editing formulas and see real-time calculation")
    
    return filename

if __name__ == '__main__':
    create_test_excel_with_formulas()